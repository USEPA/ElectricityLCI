# -*- coding: utf-8 -*-
"""
Created on Tue Jul 11 15:48:14 2017

@author: tghosh
"""

import matplotlib.pyplot as plt
import numpy as np
#import plotly.plotly as py
import openpyxl
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
import os
import statistics

print('Opening workbook...')


#We Need Sorted file for this to work Sorted acc to EIS CODE
#change the column of generation to FLoat numbers in excel.
#Should have an extra Sheet2
wb = openpyxl.load_workbook('chk4.xlsx')
chk = wb.get_sheet_by_name('Sheet1')

egridid = chk['C2':'C2219']
eisid = chk['E2':'E2219']
gen = chk['D2':'D2219']

greenFill = PatternFill(start_color='FFADFFA4',
                   end_color='FFADFFA4',
                   fill_type='solid')

sheet = wb.get_sheet_by_name('Sheet2')
i = 0;
s = 0;


for k in range(1,232):
    v = chk.cell(row = 1, column = k).value;
    sheet.cell(row = 1, column = k).value = v;


while i < 2218:
    s = eisid[i][0].value
    sum = 0;
    ar = [];
    for j in range(0,2218):
        if s == eisid[j][0].value and gen[j][0].value != None and type(gen[j][0].value) != str:
              sum = sum + float(gen[j][0].value);
              ar.append(gen[j][0].value)
               
    if(len(ar)>1):
        print(i);
        for k in range(1,7):
           v = chk.cell(row = i+2, column = k).value;
           sheet.cell(row = i+2, column = k).value = v;
        for k in range(7,232):
           v = chk.cell(row = i+2, column = k).value;
           if type(v)==float or type(v)==int:
             #print(type(gen[i][0].value)) 
             if gen[i][0].value != None:
               sheet.cell(row = i+2, column = k).value = v*gen[i][0].value/sum;
               sheet.cell(row = i+2, column = k).fill = greenFill;
           
    else:
        for k in range(1,232):
            v = chk.cell(row = i+2, column = k).value;
            sheet.cell(row = i+2, column = k).value = v;
            
            
    i = i+1;      
    
    
wb.save('FInalNeiemissions.xlsx')    

