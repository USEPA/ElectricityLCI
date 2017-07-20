# -*- coding: utf-8 -*-
"""
Created on Thu Jun 15 15:57:07 2017

@author: tghosh
"""

# -*- coding: utf-8 -*-
"""
Created on Mon Jun 12 11:52:37 2017

@author: tghosh
"""

# -*- coding: utf-8 -*-
"""
THis file is used to average the emnissions for fuel sources and different regions
PLEASE MAKE SURE YOUR EGRIDNEWCORRECT FILE IS SORTED ALPHABETICALLY with REGIONS OTHERWISE IT MIGHT NOT WORK
"""


import matplotlib.pyplot as plt
import numpy as np
#import plotly.plotly as py
import openpyxl
import os
import statistics

#os.chdir('C:/Users/tghosh/Desktop/LCI work/Day 30/New Average Egrid')
print('Opening workbook...')

#ar = [None]*3063
#wb = openpyxl.load_workbook('NOX.xlsx')

wb = openpyxl.load_workbook('NEWLCI.xlsx')
sheet1 = wb.get_sheet_by_name('Sheet1')
sheet = wb.get_sheet_by_name('Sheet2')
reg1 = sheet['A4':'A29'] 
src1 = sheet['C4':'C21'] 

wb2 = openpyxl.load_workbook('egridcorrectnew.xlsx')
sheet2 = wb2.get_sheet_by_name('less10%')
nox = sheet2['I2':'I8294'] 
so2 = sheet2['J2':'J8294'] 
co2 = sheet2['K2':'K8294'] 
ch4 = sheet2['L2':'L8294'] 
n2o = sheet2['M2':'M8294'] 

src2 = sheet2['G2':'G8294'] 
src3 = sheet2['H2':'H8294'] 
pm = sheet2['N2':'N8294'] 
reg2 = sheet2['F2':'F8294'] 


#sheet3 = wb.create_sheet('Sheet2')

for i in range(0,26):
  for k in range(0,18): 
    ar1=[];
    ar2=[];
    ar3=[];
    ar4=[];
    ar5=[];  
    for j in range(0,8293):
     if n2o[j][0].value != ' ':
      if reg1[i][0].value == reg2[j][0].value:
          if src1[k][0].value == src2[j][0].value or src1[k][0].value == src3[j][0].value:
             if nox[j][0].value!=0:  
              ar1.append(nox[j][0].value);
             if so2[j][0].value!=0:   
              ar2.append(so2[j][0].value);
             if co2[j][0].value!=0:  
              ar3.append(co2[j][0].value);
             if ch4[j][0].value!=0:   
              ar4.append(ch4[j][0].value);
             if n2o[j][0].value!=0:
              ar5.append(n2o[j][0].value);
          elif src1[k][0].value == 'SOLAR PV' and src3[j][0].value == 'SOLAR' and pm[j][0].value == 'PV':
             if nox[j][0].value!=0:  
              ar1.append(nox[j][0].value);
             if so2[j][0].value!=0:   
              ar2.append(so2[j][0].value);
             if co2[j][0].value!=0:  
              ar3.append(co2[j][0].value);
             if ch4[j][0].value!=0:   
              ar4.append(ch4[j][0].value);
             if n2o[j][0].value!=0:   
              ar5.append(n2o[j][0].value);
          elif src1[k][0].value == 'SOLAR THERMAL' and src3[j][0].value == 'SOLAR' and pm[j][0].value != 'PV':
             if nox[j][0].value!=0:  
              ar1.append(nox[j][0].value);
             if so2[j][0].value!=0:   
              ar2.append(so2[j][0].value);
             if co2[j][0].value!=0:  
              ar3.append(co2[j][0].value);
             if ch4[j][0].value!=0:   
              ar4.append(ch4[j][0].value);
             if n2o[j][0].value!=0:   
              ar5.append(n2o[j][0].value);
          elif src1[k][0].value == 'GEOTHERMAL BT' and src2[j][0].value == 'GEO' and pm[j][0].value == 'BT':
             if nox[j][0].value!=0:  
              ar1.append(nox[j][0].value);
             if so2[j][0].value!=0:   
              ar2.append(so2[j][0].value);
             if co2[j][0].value!=0:  
              ar3.append(co2[j][0].value);
             if ch4[j][0].value!=0:   
              ar4.append(ch4[j][0].value);
             if n2o[j][0].value!=0:   
              ar5.append(n2o[j][0].value);
          elif src1[k][0].value == 'GEOTHERMAL FT' and src2[j][0].value == 'GEO' and pm[j][0].value != 'BT':
             if nox[j][0].value!=0:  
              ar1.append(nox[j][0].value);
             if so2[j][0].value!=0:   
              ar2.append(so2[j][0].value);
             if co2[j][0].value!=0:  
              ar3.append(co2[j][0].value);
             if ch4[j][0].value!=0:   
              ar4.append(ch4[j][0].value);
             if n2o[j][0].value!=0 and n2o[j][0].value!=None:   
              ar5.append(n2o[j][0].value);
    
    
    
    if(len(ar1)>0):        
     sheet1.cell(row = i*18+k+3,column = 3).value = np.mean(ar1)
    else:
     sheet1.cell(row = i*18+k+3,column = 3).value = 'NA'
    if(len(ar2)>0): 
     sheet1.cell(row = i*18+k+3,column = 4).value = np.mean(ar2)
    else:
     sheet1.cell(row = i*18+k+3,column = 4).value = 'NA'
    if(len(ar3)>0): 
     sheet1.cell(row = i*18+k+3,column = 5).value = np.mean(ar3)
    else:
     sheet1.cell(row = i*18+k+3,column = 5).value = 'NA'
    if(len(ar4)>0): 
     sheet1.cell(row = i*18+k+3,column = 6).value = np.mean(ar4)
    else:
     sheet1.cell(row = i*18+k+3,column = 6).value = 'NA'
    if(len(ar5)>0):
     sheet1.cell(row = i*18+k+3,column = 7).value = np.mean(ar5)
    else:
     sheet1.cell(row = i*18+k+3,column = 7).value = 'NA'



wb.save('chk.xlsx')



#############################################################################################################################################







