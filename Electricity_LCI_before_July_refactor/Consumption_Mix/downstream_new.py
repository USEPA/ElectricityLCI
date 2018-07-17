import pandas as pd
from copy import copy, deepcopy
import openpyxl
import numpy as np
import os
import sys


data_dir = os.path.dirname(os.path.realpath(__file__))
data_dir = data_dir+'//Output//';

def egrid_func(a,b):

        #Reading the facility file
        egrid1 = pd.read_csv("egrid_2014_1.csv", header=0, error_bad_lines=False)
        
        
        #Reading the flow by facility file
        egrid2 = pd.read_csv("egrid2014_trircrainfoneiegrid.csv", header=0, error_bad_lines=False)
        
        egrid2 = egrid2[egrid2['Source'] == 'eGRID']
        
        egrid2 = egrid2.drop_duplicates(keep = 'first')
        
        
        egrid3 = pd.pivot_table(egrid2,index = 'FacilityID' ,columns = 'FlowName', values = 'FlowAmount').reset_index()
        
        
        #Merging dataframes together
        egrid = egrid1.merge(egrid3, left_on = 'FacilityID', right_on = 'FacilityID')
        egrid[['Heat input','Net generation']] = egrid[['Heat input','Net generation']].apply(pd.to_numeric,errors = 'coerce')
        
        
        
        #Calculating efficiency
        egrid.loc[:,"Efficiency"] = egrid['Net generation'].values*100/egrid['Heat input'].values
        egrid = egrid.sort_values(by = ['Efficiency'], axis=0, ascending=True, inplace=False, kind='quicksort', na_position='last')
       
        #Replace inf
        egrid = egrid.replace([np.inf, -np.inf], np.nan)
        #Replacing all the facilities with no net generation or Heat input reported
        egrid = egrid.dropna(subset = ['Efficiency'])
       
        #Dropping less than 10%
        egrid_new1 = egrid[(egrid['Efficiency'] >= a) & (egrid['Efficiency'] <= b)]
        
        
        
       
        
        col = list(egrid_new1['Heat input'])        
        egrid_new1['HeatInput'] = col;
        cols = egrid_new1.columns.tolist()
        egrid_new1 = egrid_new1[[cols[-1]] + cols[:-1]] 
        
        
        col = list(egrid_new1['Net generation'])         
        egrid_new1['NetGen'] = col;
        cols = egrid_new1.columns.tolist()
        egrid_new1 = egrid_new1[[cols[-1]] + cols[:-1]] 
        egrid_new1 = egrid_new1.drop(columns = ['Net generation'])       
        egrid_new1 = egrid_new1.drop(['Heat input','Efficiency'],axis = 1)
        egrid_new1['NetGen'] = egrid_new1['NetGen']*0.277778/1000
        #egrid_new1.to_csv('chk.csv')
        return egrid_new1

def surplus_pool():
    
            wb2 = openpyxl.load_workbook('eGRID_Consumption_Mix.xlsx',data_only=True)
            data = wb2.get_sheet_by_name('ConsumptionMixContribution') 
            
            reg1 = data['H4':'H13']
            #trade = data['F4':'F29']
            
            matrix = data['I3':'AP13']
            
            
            for i in range(0,10):
                        chk = 0;
                        Reg = reg1[i][0].value
                        wb = openpyxl.load_workbook('template.xlsx')
                        gi = wb.get_sheet_by_name('General information') 
                        io = wb.get_sheet_by_name('InputsOutputs')
                        
                        
                        global blank_row;
                                   
                        blank_row = 6;
                        Column_number = 43 #THis is based on the template number of columns in the template. 
                           
                        
                        index = 2; #Starting the index at a specific point
                        #REading the list of fuel for the entered region
                        
                        
                                       
                        
                        #This block is used for filling up the General Infomration Sheet
                        
                        
                                
                        
                        cn = 3;
                        
                        gi['D4'].value = 'Electricity '
                        
                        gi['D5'].value = 'from Surplus Pool Mix';
                        
                        gi['D6'].value = 'at region '+ Reg;
                        
                        gi['D7'].value = 'Consumption Mix';
                        
                        gi['D10'].value = 'Consumption Mix Electricity from Surplus in the '+Reg+' region'
                        
                        
                        gi['D11'].value = '22: Utilities/2211: Electric Power Generation, Transmission and Distribution/'
                        
                        gi['D12'].value = 'FALSE'
                        
                        gi['D14'].value = 'Electricity; voltage?'
                        
                        gi['D16'].value = '01/01/2017'
                        
                        gi['D17'].value = '12/31/2017'
                        
                        gi['D18'].value = '2014'
                        
                        gi['D20'].value = 'US-NERC-'+Reg
                        
                        v= '' ;

                        
                        gi['D24'].value = 'EGRID subregion definitions:<https://www.epa.gov/energy/emissions-generation-resource-integrated-database-egrid>\nNERC region: '+v+'\nStates totally or partially included: <autofill>'
                        
                  
                         
                        
                        
                        #linestring = open('out.txt', 'r').read()
                        
                        gi['D26'].value = 'This is the Surplus Pool.'
                        
                        filename = 'Surplus '+ Reg + '.xlsx'
                        
                        
                        
                        #This block is used for filling up the InputOutput Sheet ONLY EMISSIONS
                        
                        
                        #assuming the starting row to fill in the emissions is the 5th row Blank
                        
                        blank_row = 6;
                        
                        #Reading All the files
                        io['D5'].value =  'Electricity; from surplus pool mix; at region '+Reg  
                        io['E5'].value = '22: Utilities/2211: Electric Power Generation, Transmission and Distribution/'
                                                   
                        #This function is necessary to creates new line in the template for writing input and output flows. 
                        def createblnkrow(br):
                            for i in range(1,Column_number):
                              
                              v = io.cell(row = br,column = i).value
                              io.cell(row = br+1 ,column = i).value = v
                              io.cell(row = br+1 ,column = i).font = copy(io.cell(row = br,column = i).font)
                              io.cell(row = br+1 ,column = i).border = copy(io.cell(row = br,column = i).border)
                              io.cell(row = br+1 ,column = i).fill = copy(io.cell(row = br,column = i).fill)
                            br = br + 1;
                              #print(io.cell(row = 6,column = 4).value);
                            return br
                        
                        row = blank_row;
                        col = 0;
                         
                        index = 2;
                                    
                        
                        
                           
                              #if trade[i][0].value != 0:
                        for j in range(0,7):
                                       
                                       if matrix[i+1][j].value != None and matrix[i+1][j].value !=0:
                                                   #print(matrix[i+1][j].value)
                                                   blank_row = createblnkrow(blank_row)
                                                   io[row][0].value = index;
                                                   io[row][1].value = 5;
                                                   io[row][3].value = 'Canada Provinces - '+matrix[0][j].value+' electricity';
                                                   io[row][4].value = '22: Utilities/2211: Electric Power Generation, Transmission and Distribution/'
                                                   io[row][6].value =  matrix[i+1][j].value
                                                   io[row][7].value = 'MWh';
                                                   io[row][21].value = 'Electricity Trade';                           
                                                   row = row+1;
                                                   index = index+1;
                                                   chk = 1;
                                                   
                        for j in range(7,8):
                                       
                                        if matrix[i+1][j].value != None and matrix[i+1][j].value !=0:
                                                   #print(matrix[i+1][j].value)
                                                   blank_row = createblnkrow(blank_row)
                                                   io[row][0].value = index;
                                                   io[row][1].value = 5;
                                                   io[row][3].value = matrix[0][j].value+'- electricity';
                                                   io[row][4].value = '22: Utilities/2211: Electric Power Generation, Transmission and Distribution/'
                                                   io[row][6].value =  matrix[i+1][j].value
                                                   io[row][7].value = 'MWh';
                                                   io[row][21].value = 'Electricity Trade';                           
                                                   row = row+1;
                                                   index = index+1; 
                                                   chk = 1;
                                                   
                        for j in range(8,34):
                                       
                                        if matrix[i+1][j].value != None and matrix[i+1][j].value !=0:
                                                   #print(matrix[i+1][j].value)
                                                   blank_row = createblnkrow(blank_row)
                                                   io[row][0].value = index;
                                                   io[row][1].value = 5;
                                                   io[row][3].value = 'Electricity 100 - 120V';
                                                   io[row][5].value = matrix[0][j].value
                                                   io[row][4].value = '22: Utilities/2211: Electric Power Generation, Transmission and Distribution/'
                                                   io[row][6].value =  matrix[i+1][j].value
                                                   io[row][7].value = 'MWh';
                                                   io[row][21].value = 'Electricity Trade';                           
                                                   row = row+1;
                                                   index = index+1;
                                                   chk = 1;
                        #This function  is to find all other output flows and input fuels.
                        
                             
                                    
                                    
                        if chk == 1:            
                          wb.save(data_dir+filename)
                        
                        
                        
            print('Compilation Successful')       
            

def consumption_mix():

            wb2 = openpyxl.load_workbook('eGRID_Consumption_Mix.xlsx',data_only=True)
            data = wb2.get_sheet_by_name('ConsumptionMixContribution') 
            
            reg1 = data['H4':'H13']
            trade = data['F4':'F29']
            reg2 = data['C4':'C29']
            reg3 = data['A4':'A29']
            matrix = data['I3':'AP13']
            gen = data['E4':'E29']
            
            for k in range(0,26):
                        chk = 0;
                        Reg = reg2[k][0].value
                        wb = openpyxl.load_workbook('template.xlsx')
                        gi = wb.get_sheet_by_name('General information') 
                        io = wb.get_sheet_by_name('InputsOutputs')
                        
                        
                        global blank_row;
                                   
                        blank_row = 6;
                        Column_number = 43 #THis is based on the template number of columns in the template. 
                           
                        
                        index = 2; #Starting the index at a specific point
                        #REading the list of fuel for the entered region
                        
                        
                                       
                        
                        #This block is used for filling up the General Infomration Sheet
                        
                        
                                
                        
                        cn = 3;
                        
                        gi['D4'].value = 'Electricity '
                        
                        gi['D5'].value = 'from Consumption Mix';
                        
                        gi['D6'].value = 'at region '+ Reg;
                        
                        gi['D7'].value = 'Consumption Mix';
                        
                        gi['D10'].value = 'Consumption mix electricity in the '+Reg+' region'
                        
                        
                        gi['D11'].value = '22: Utilities/2211: Electric Power Generation, Transmission and Distribution/'
                        
                        gi['D12'].value = 'FALSE'
                        
                        gi['D14'].value = 'Electricity; voltage?'
                        
                        gi['D16'].value = '01/01/2017'
                        
                        gi['D17'].value = '12/31/2017'
                        
                        gi['D18'].value = '2014'
                        
                        gi['D20'].value = 'US-eGRID-'+Reg
                        
                        v= '' ;
                        #THis function is used for finding the NERC regions. THese input files are built manually. 
            
                        
                        gi['D24'].value = 'EGRID subregion definitions:<https://www.epa.gov/energy/emissions-generation-resource-integrated-database-egrid>\nNERC region: '+v+'\nStates totally or partially included: <autofill>'
                        
                        #This is for printing the Prime Movers. Written to a output file and reread back. 
                        #def pm(fn,rg):
                        #    f = open('out.txt', 'w')
                        #    for row in primemover.itertuples():
                        ##      if row[2] == rg:
                         #       if row[3]  == fn or row[4] == fn or row[5] == fn or row[6] == fn:
                        #          if row[7] != None:                   
                        #             f.write(row[7])
                        #             f.write(',')
                        #    f.close()        
                        
                         
                        
                        
                        #linestring = open('out.txt', 'r').read()
                        
                        gi['D26'].value = 'This is the Consumption Mix.'
                        
                        filename = 'Consumption Mix '+ Reg + '.xlsx'
                        
                        
                        
                        #This block is used for filling up the InputOutput Sheet ONLY EMISSIONS
                        
                        
                        #assuming the starting row to fill in the emissions is the 5th row Blank
                        
                        blank_row = 6;
                        
                        #Reading All the files
                        io['D5'].value =  Reg+' Consumption Mix electricity';   
                        
                        #This function is necessary to creates new line in the template for writing input and output flows. 
                        def createblnkrow(br):
                            for i in range(1,Column_number):
                              
                              v = io.cell(row = br,column = i).value
                              io.cell(row = br+1 ,column = i).value = v
                              io.cell(row = br+1 ,column = i).font = copy(io.cell(row = br,column = i).font)
                              io.cell(row = br+1 ,column = i).border = copy(io.cell(row = br,column = i).border)
                              io.cell(row = br+1 ,column = i).fill = copy(io.cell(row = br,column = i).fill)
                            br = br + 1;
                              #print(io.cell(row = 6,column = 4).value);
                            return br
                        
                        row = blank_row;
                        col = 0;
                         
                        index = 2;
            #Filling up the generation Mix flow at first                        
                        blank_row = createblnkrow(blank_row)
                        io[row][0].value = index;
                        io[row][1].value = 5;
                        io[row][3].value = 'Electricity 100 - 120V';
                        io[row][4].value = '22: Utilities/2211: Electric Power Generation, Transmission and Distribution/'
                        io[row][5].value = reg2[k][0].value;
                        io[row][6].value = gen[k][0].value
                        io[row][7].value = 'MWh';
                        io[row][21].value = 'Electricity Generation Mix';                           
                        row = row+1;
                        index = index+1;
                        
                            
                        for i in range(0,10):
                                 
                                if reg3[k][0].value == reg1[i][0].value:
                                   temp = trade[k][0].value
                                   
                                   if trade[k][0].value != 0:
                                        for j in range(0,34):
                                                       
                                                       if matrix[i+1][j].value != None and matrix[i+1][j].value !=0:
                                                                   #print(matrix[i+1][j].value)
                                                                   blank_row = createblnkrow(blank_row)
                                                                   io[row][0].value = index;
                                                                   io[row][1].value = 5;
                                                                   io[row][3].value = 'Electricity; from surplus pool mix; at region '+reg3[k][0].value;
                                                                   io[row][4].value = '22: Utilities/2211: Electric Power Generation, Transmission and Distribution/'
                                                                   io[row][6].value = trade[k][0].value 
                                                                   io[row][7].value = 'MWh';
                                                                   io[row][21].value = 'Electricity Trade';                           
                                                                   row = row+1;
                                                                   index = index+1;
                                                                   chk = 1;
                                                                   break;                                
             #This part is used for the cases of FRCC where even though there is a trade in value, the value is overwritten as 0 because of no values in the grey box.                                                       
                                        if chk != 1:
                                             # WHy is the row 6? because the generation flow which is being overwrtten is always at row 6 because its the first thing that is written.       
                                                  
                                             io[6][6].value = 1.0
                                         
                                                                   
                        
                                        
                             
                                                   
                        wb.save(data_dir+filename)
                        
                        
                        
            print('Compilation Successful')  


surplus_pool()
#surplus_pool()