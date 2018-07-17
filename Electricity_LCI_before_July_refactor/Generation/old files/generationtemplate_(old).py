#Writes out generation processes. Not currently working because relies on inventory from old sources

import pandas as pd
from copy import copy, deepcopy
import openpyxl
import numpy as np
import os
import sys

def egrid_func(a,b):

        #Reading the facility file
        egrid1 = pd.read_csv("egrid_2014_1.csv", header=0, error_bad_lines=False)
        
        
        #Reading the flow by facility file
        egrid2 = pd.read_csv("egrid_2014.csv", header=0, error_bad_lines=False)
        
        
        egrid3 = egrid2.pivot(index = 'FacilityID',columns = 'FlowName', values = 'FlowAmount').reset_index()
        
        
        #Merging dataframes together
        egrid = egrid1.merge(egrid3, left_on = 'FacilityID', right_on = 'FacilityID')
        egrid[['Heat input','Net generation']] = egrid[['Heat input','Net generation']].apply(pd.to_numeric,errors = 'coerce')
        
        
        
        #Calculating efficiency
        egrid.loc[:,"Efficiency"] = egrid['Net generation'].values*100/egrid['Heat input'].values
        egrid = egrid.sort_values(by = ['Efficiency'], axis=0, ascending=True, inplace=False, kind='quicksort', na_position='last')
       
        #Replace inf
        egrid = egrid.replace([np.inf, -np.inf], np.nan)
        #Replacing all the facilities with no net generation or Heat input reported
        egrid = egrid.dropna(subset = ['Efficiency'])
       
        #Dropping less than 10%
        egrid_new1 = egrid[(egrid['Efficiency'] >= a) & (egrid['Efficiency'] <= b)]
        
        egrid_new1.to_csv('chk.csv')
        
       
        
        col = list(egrid_new1['Heat input'])        
        egrid_new1['HeatInput'] = col;
        cols = egrid_new1.columns.tolist()
        egrid_new1 = egrid_new1[[cols[-1]] + cols[:-1]] 
        
        
        col = list(egrid_new1['Net generation'])         
        egrid_new1['NetGen'] = col;
        cols = egrid_new1.columns.tolist()
        egrid_new1 = egrid_new1[[cols[-1]] + cols[:-1]] 
        egrid_new1 = egrid_new1.drop(columns = ['Net generation'])       
        egrid_new1 = egrid_new1.drop(['Heat input','Efficiency'],axis = 1)
        egrid_new1['NetGen'] = egrid_new1['NetGen']*0.277778/1000
        
        return egrid_new1
    
def compilation(db):
        #Troy Method
        #Creating copy of database by substitution the NA emissions with zero
        db1 = db.fillna(value = 0)
        
        #Removing all rows here emissions are not reported for second dataframe
        #db2 = db.dropna(how = db[db.columns[1]])
        db2 = db.dropna()
        
        
        #keeping the unreported emissions and facilities in separate database
        #db3 = db[pd.isnull(db.iloc[:,1])]
        
        if db2.empty == True:
            ef1 = np.sum(db1.iloc[:,1])/np.sum(db1.iloc[:,0])
            return ef1
    
        ef1 = np.sum(db1.iloc[:,1])/np.sum(db1.iloc[:,0])
        
        ef2 = np.sum(db2.iloc[:,1])/np.sum(db2.iloc[:,0])
        
        weight = np.sum(db2.iloc[:,0])/np.sum(db1.iloc[:,0])
        final_ef = ef2*weight + (1-weight)*ef1
        return final_ef


def tri_func(egrid):
        
        #READING tri database
        stewi_combo = pd.read_csv("tri_nei_rcrainfo_combined_for_egrid.csv", header=0, error_bad_lines=False)  
        
        #Droppping duplicates in STEWI
        stewi_combo = stewi_combo.drop_duplicates(keep = 'first')
        stewi_combo = stewi_combo[['eGRID_ID','FlowName','Compartment','FlowAmount','Source','ReliabilityScore']]
        stewi_combo2 = stewi_combo.groupby(['eGRID_ID','FlowName','Compartment','Source','ReliabilityScore'])['FlowAmount'].sum()
        stewi_combo2 = stewi_combo2.reset_index()
        
        #stewi_combo2 = stewi_combo2.dropna(axis = 0, how = 'any')
        #stewi_combo reshaping pivot not working so using pivot table
        stewi_combo3 = pd.pivot_table(stewi_combo2,index = ['eGRID_ID','Compartment','Source','ReliabilityScore'], columns = 'FlowName', values = 'FlowAmount')
        
        stewi_combo3 = stewi_combo3.reset_index()

        egrid = egrid[['NetGen','FacilityID']]
                
        #Merging egrid and TRI
        database = egrid.merge(stewi_combo3,left_on ='FacilityID', right_on = 'eGRID_ID')
        
        database = database.drop(columns = ['FacilityID','eGRID_ID'])
        return database              
       #database.to_csv('chk.csv')
                
    
    
    
    
    
#This function is necessary to creates new line in the template for writing input and output flows. 
def createblnkrow(br,io):
    
    Column_number = 43 #THis is based on the template number of columns in the template. 
    for i in range(1,Column_number):
      
      v = io.cell(row = br,column = i).value
      io.cell(row = br+1 ,column = i).value = v
      io.cell(row = br+1 ,column = i).font = copy(io.cell(row = br,column = i).font)
      io.cell(row = br+1 ,column = i).border = copy(io.cell(row = br,column = i).border)
      io.cell(row = br+1 ,column = i).fill = copy(io.cell(row = br,column = i).fill)
    
    br = br + 1;
      #print(io.cell(row = 6,column = 4).value);
    return br


    
def generator(a,b,Reg):
            
            global year; 
            year = '2016'
            database = egrid_func(a,b)
            database = database[database['eGRID subregion acronym'] == Reg]
            d_list = ['eGRID','NEI','TRI','RCRAInfo']
                            
            
            fuel_name = pd.read_csv("fuelname.csv", header=0, error_bad_lines=False)
            for row in fuel_name.itertuples():
              #assuming the starting row to fill in the emissions is the 5th row Blank
              
              global blank_row;    
              blank_row = 6;
              index = 2;
              for roww in database.itertuples():
                
                
                #The fuels and their types are in two different columns
                if row[1] == roww[9] or row[1] == roww[10]:              
                    
                    fuelname = row[2]
                    fuelheat = row[3]
                    print(fuelname)
                                      
        
                    #Read the template files.                         
                    wbook = openpyxl.load_workbook('Fed_template_Main_19Dec2016.xlsx')
                    gi = wbook['General information']
                    
                    #This block is used for filling up the General Infomration Sheet      
                    gi['D4'].value = 'Electricity'
                    
                    gi['D5'].value = 'from '+ fuelname;
                    
                    gi['D6'].value = 'at generating facility';
                    
                    #Mix type
                    #gi['D7'].value = 'Production Mix'; 
                    
                    
                    
                    gi['D10'].value = 'Electricity from '+ fuelname+' using power plants in the '+Reg+' region'
                    
                    #This function is used for the name of the fuels.
                    #def pm1(fn):
                    #    for row in primemover.itertuples():
                    #       if row[3]  == fn or row[4] == fn or row[5] == fn or row[6] == fn:
                    #          if row[4] != None:                   
                    #             return row[4].capitalize()
                    
                    gi['D11'].value = '22: Utilities/2211: Electric Power Generation, Transmission and Distribution/'+str(fuelname)
                    
                    gi['D12'].value = 'FALSE'
                    
                    gi['D14'].value = 'Electricity; voltage?'
                    
                    gi['D16'].value = '01/01/2017'
                    
                    gi['D17'].value = '12/31/2017'
                    
                    gi['D18'].value = year
                    
                    gi['D20'].value = 'US-eGRID-'+Reg
                
                    gi['D21'].value = 'F'
                    v= [];
                    v1= [];
                    '''
                    #THis function is used for finding the NERC regions. THese input files are built manually. 
                    for row in temp.itertuples():
                        if row[2] == Reg:
                           v = row[1]
                           
                    for row in state.itertuples():
                        if row[1] == Reg:
                            v1 = row[2];
                    
                    gi['D24'].value = 'database subregion definitions:<https://www.epa.gov/energy/emissions-generation-resource-integrated-database-database>\nNERC region: '+v+'\nStates totally or partially included: '+v1
                    '''
                    
                    for rowww in database.itertuples():
                        if row[1] == roww[9] or row[1] == roww[10]:
                            v1.append(rowww[4])
                            v.append(roww[11])
                            
                    v2 = list(set(v1)) 
                    
                    str1 = ','.join(v2)
                    
                    v2 = list(set(v))
                    str2 = ','.join(v2)
                    gi['D24'].value = 'database subregion definitions:<https://www.epa.gov/energy/emissions-generation-resource-integrated-database-database>\nNERC region: '+str2+'\nStates totally or partially included: '+str1
                   
        
                    '''
                    #This is for printing the Prime Movers. Written to a output file and reread back. 
                    def pm(fn,rg):
                        f = open('out.txt', 'w')
                        for row in primemover.itertuples():
                          if row[2] == rg:
                            if row[3]  == fn or row[4] == fn or row[5] == fn or row[6] == fn:
                              if row[7] != None:                   
                                 f.write(row[7])
                                 f.write(',')
                        f.close()        
                    
                     
                    pm(Fuel,Reg)
                    
                    #linestring = open('out.txt', 'r').read()
                    
                    gi['D26'].value = 'This is an aggregation of technology types within this eGRID subregion.  List of prime movers:'+linestring
                    
                    '''
                    
               
                    #croppping the database according to the current fuel being considered
                    database_f1 = database[database['Plant primary coal/oil/gas/ other fossil fuel category'] == row[1]]
                    if database_f1.empty == True:
                          database_f1 = database[database['Plant primary fuel'] == row[1]]              
                    
                    
                  
                    #This block is used for filling up the InputOutput Sheet ONLY EMISSIONS
                    io = wbook['InputsOutputs']
                    
                    

                    #Fillin up with reference only only
                    io['A5'].value = 1
                    io['C5'].value =  0;
                    io['D5'].value =  'Electricity 100 - 120V';
                    io['E5'].value = '22: Utilities/2211: Electric Power Generation, Transmission and Distribution/'+str(fuelname)
                    io['G5'].value =  1 
                    io['H5'].value =  'MWh'
                    row1 = blank_row;
                    
                    blank_row = createblnkrow(blank_row,io)
                    
                    
                    if database_f1['HeatInput'].mean() != 0 and fuelheat != 0:
                    #Fillin up the fuel flow
                        io[row1][0].value = index;
                        io[row1][1].value = 5;
                        io[row1][3].value = fuelname;
                        io[row1][4].value = 'Input Fuel'
                        io[row1][6].value = np.sum(database_f1['HeatInput'])/np.sum(database_f1['NetGen']);
                        io[row1][7].value = 'kg';
                        io[row1][21].value = 'database data with plants over 10% efficiency';
                        io[row1][26].value = 'Normal';
                        io[row1][27].value = np.sum(database_f1['HeatInput'])/np.sum(database_f1['NetGen']);
                        io[row1][28].value = database_f1['HeatInput'].std();
                        io[row1][29].value = database_f1['HeatInput'].min();
                        io[row1][30].value = database_f1['HeatInput'].max();
                        io[row1][33].value = 'eGRID '+year;
                        row1 = blank_row;
                        index = index + 1;
                    
                    
                    
                    for x in d_list:
                        if x == 'eGRID': 
                            database_f3 = database_f1[['NetGen','Carbon dioxide', 'Nitrous oxide', 'Nitrogen oxides', 'Sulfur dioxide', 'Methane']]
                            (blank_row,row1,index) = flowwriter(database_f3,io,row1,index,blank_row,'air',x)
                        
                        elif x != 'eGRID':
                            database_f3 = tri_func(database_f1);
                            if x == 'TRI':                     
                                database_f3 = database_f3[database_f3['Source']=='TRI']
                            elif x == 'NEI':
                                database_f3 = database_f3[database_f3['Source']=='NEI']
                            elif x == 'RCRAInfo':
                                database_f3 = database_f3[database_f3['Source']=='RCRAInfo']
                            
                            if database_f3.empty != True:
                                                      
                                #water
                                d1 = database_f3[database_f3['Compartment']=='air']
                                d1 = d1.drop(columns = ['Compartment','Source'])
                                
                                if d1.empty != True:
                                  (blank_row,row1,index) = flowwriter(d1,io,row1,index,blank_row,'air',x)                            
                                
                                
                                #water
                                d1 = database_f3[database_f3['Compartment']=='water']
                                d1 = d1.drop(columns = ['Compartment','Source'])
                                
                                if d1.empty != True:
                                  (blank_row,row1,index) = flowwriter(d1,io,row1,index,blank_row,'water',x)
                                
                                #soil
                                d1 = database_f3[database_f3['Compartment']=='soil']
                                d1 = d1.drop(columns = ['Compartment','Source'])
                                
                                if d1.empty != True:
                                  (blank_row,row1,index) = flowwriter(d1,io,row1,index,blank_row,'soil',x)
                                
                                
                                
                                #waste
                                d1 = database_f3[database_f3['Compartment']=='waste']
                                d1 = d1.drop(columns = ['Compartment','Source'])
                                
                                if d1.empty != True:
                                  (blank_row,row1,index) = flowwriter(d1,io,row1,index,blank_row,'waste',x)

                             
                        
                            
                            
                            
                            

                    filename = fuelname+"_"+Reg+".xlsx"
                    data_dir = os.path.dirname(os.path.realpath(__file__))+"\\results\\"
                    wbook.save(data_dir+filename)
                    break;


def flowwriter(database_f1,io,row1,index,blank_row,comp,y):
    
    flownames = list(database_f1.columns.values) 
    for i in database_f1.iteritems():
           
      if str(i[0]) != 'NetGen' and str(i[0]) != 'ReliabilityScore': 
        database_f2 = database_f1[['NetGen',i[0]]] 
        if(compilation(database_f2) != 0 and compilation(database_f2)!= None):
            blank_row = createblnkrow(blank_row,io)
            io[row1][0].value = index;
            io[row1][2].value = 4;
            io[row1][3].value = i[0];
            io[row1][4].value = 'Elementary Flows/'+str(comp)+'/unspecified'
            io[row1][6].value = compilation(database_f2);
            io[row1][7].value = 'kg';
            io[row1][21].value = 'database data with plants over 10% efficiency';
            io[row1][26].value = 'Normal';
            io[row1][27].value = compilation(database_f2);
            io[row1][28].value = database_f1[i[0]].std();
            io[row1][29].value = database_f1[i[0]].min();
            io[row1][30].value = database_f1[i[0]].max();
            io[row1][33].value = y+' '+year;
            print(i[0])
            row1 = row1+1;
            index = index+1;
            
    return (blank_row,row1,index)
#egrid = egrid_func(10,100)
#p = tri_func(egrid)
region = pd.read_excel('Region.xlsx')
for ind in region.itertuples():
    print(ind[1])
    
    generator(10,100,str(ind[1]))     
