#Generates generation mixes for each eGRID subregion

import pandas as pd
from copy import copy
import openpyxl

import globals
datadir = globals.datadir
outputdir = globals.outputdir


egrid = pd.read_csv(datadir+"eGRID_2014.csv", header=0, error_bad_lines=False)
temp = egrid[['NERC region acronym',' eGRID subregion acronym ']]
temp = temp.drop_duplicates()

primemover = pd.read_csv("prime_moverlist.csv", header=0, error_bad_lines=False)
fuellist = pd.read_csv("fuel_list.csv", header=0, error_bad_lines=False)
fuel_name = pd.read_csv("fuelname.csv", header=0, error_bad_lines=False)
fuellist = fuellist.dropna(axis = 0, subset = ['FuelList'])
fuellist = fuellist.dropna(axis = 0, subset = [' eGRID subregion acronym '])



#THis is for inputting the region name from the user for which the templates will be built. 

Reg = input('Please enter 4 lettered eGRID subregion here in uppercase: ')
print('Let the magic begin ---- ')
#Read the template files. 

wb = openpyxl.load_workbook(datadir+'template.xlsx')
gi = wb['General information']
io = wb['InputsOutputs']
global blank_row;
           
blank_row = 6;
Column_number = 43 #THis is based on the template number of columns in the template. 
   

index = 2; #Starting the index at a specific point
#REading the list of fuel for the entered region
for row in fuellist.itertuples():
      if row[2] == Reg:
            Fuel = row[3]
            #Fuel name is a file built manually to replace the Coded names of the fuels to full form names. CHeck the full forms please.May be wrong. 
            #Reg5 = input('Please enter region here: ') 
            for  roww in fuel_name.itertuples():
                if roww[1] == Fuel:
                    fuelname = roww[2]
                    print(fuelname)
                    break;
                    

               
            
            #This block is used for filling up the General Infomration Sheet
            
            
                    
            
            cn = 3;
            
            gi['D4'].value = 'Electricity '
            
            gi['D5'].value = 'from Generation Mix';
            
            gi['D6'].value = 'at region '+ Reg;
            
            gi['D7'].value = 'Production Mix';
            
            gi['D10'].value = 'Electricity from generation mix using power plants in the '+Reg+' region'
            
            #This function is used for the name of the fuels.
            def pm1(fn):
                for row in primemover.itertuples():
                   if row[3]  == fn or row[4] == fn or row[5] == fn or row[6] == fn:
                      if row[4] != None:                   
                         return row[4].capitalize()
            
            gi['D11'].value = '22: Utilities/2211: Electric Power Generation, Transmission and Distribution/'
            
            gi['D12'].value = 'FALSE'
            
            gi['D14'].value = 'Electricity; voltage?'
            
            gi['D16'].value = '01/01/2017'
            
            gi['D17'].value = '12/31/2017'
            
            gi['D18'].value = '2014'
            
            gi['D20'].value = 'US-eGRID-'+Reg
        
            v= '' ;
            #THis function is used for finding the NERC regions. THese input files are built manually. 
            for row in temp.itertuples():
                if row[2] == Reg:
                   v = row[1]
            
            gi['D24'].value = 'EGRID subregion definitions:<https://www.epa.gov/energy/emissions-generation-resource-integrated-database-egrid>\nNERC region: '+v+'\nStates totally or partially included: <autofill>'
            
            #This is for printing the Prime Movers. Written to a output file and reread back. 
            def pm(fn,rg):
                f = open('out.txt', 'w')
                for row in primemover.itertuples():
                  if row[2] == rg:
                    if row[3]  == fn or row[4] == fn or row[5] == fn or row[6] == fn:
                      if row[7] != None:                   
                         f.write(row[7])
                         f.write(',')
                f.close()        
            
             
            pm(Fuel,Reg)
            
            #linestring = open('out.txt', 'r').read()
            
            gi['D26'].value = 'This is an aggregation of technology types within this EGRID subregion.'
            
            filename = outputdir+'Mix '+ Reg + '.xlsx'
            
            
            
            #This block is used for filling up the InputOutput Sheet ONLY EMISSIONS
            
            
            #assuming the starting row to fill in the emissions is the 5th row Blank
            
            
            
            #Reading All the files
            io['D5'].value =  'Electricity from Generation Mix at '+Reg;   
            inp1 = pd.read_csv(datadir+"Aggregated_LCI1.csv", header=0, error_bad_lines=False)
            inp2 = pd.read_csv(datadir+"Aggregated_LCI2.csv", header=0, error_bad_lines=False)
            inp3 = pd.read_csv(datadir+"Aggregated_LCI3.csv", header=0, error_bad_lines=False)
            inp4 = pd.read_csv(datadir+"Aggregated_LCI4.csv", header=0, error_bad_lines=False)
            
            #This function is necessary to creates new line in the template for writing input and output flows. 
            def createblnkrow(br):
                for i in range(1,Column_number):
                  
                  v = io.cell(row = br,column = i).value
                  io.cell(row = br+1 ,column = i).value = v
                  io.cell(row = br+1 ,column = i).font = copy(io.cell(row = br,column = i).font)
                  io.cell(row = br+1 ,column = i).border = copy(io.cell(row = br,column = i).border)
                  io.cell(row = br+1 ,column = i).fill = copy(io.cell(row = br,column = i).fill)
                br = br + 1;
                  #print(io.cell(row = 6,column = 4).value);
                return br

            row = blank_row;
            col = 0;
         
            
            #This function is to find standard deviation values. 
            def standdev(d1,em):
                for r in d1.itertuples():
                     if r[3] == Fuel and em == r[4]: 
                       if r[2] == Reg and pd.isnull(r[5]) == False and r[5] != 0:
                            return r[5];
                        
            
            
            
            
            #This function  is to find all other output flows and input fuels.
            def main(d1,row):
                 global index
                 global blank_row;
                 for r in d1.itertuples():
                     if r[3] == Fuel: 
                       if r[2] == Reg and pd.isnull(r[6]) == False:
                           
                           blank_row = createblnkrow(blank_row)
                           io[row][0].value = index;
                           io[row][2].value = 5;
                           io[row][3].value = 'Electricity; from '+fuelname+';at generation facility';
                           io[row][4].value = '22: Utilities/2211: Electric Power Generation, Transmission and Distribution/'+fuelname
                           io[row][5].value = 'US-eGRID-'+Reg
                           io[row][6].value = r[6];
                           io[row][7].value = 'MWh';
                           io[row][21].value = 'Electricity Input';                           
                           row = row+1;
                           index = index+1; 
                           

                
           
            
            
            #Separate calls are made for Solar and Geothermal Fuel sources because of different databases. 
            #Different databases are sent because many information have to be written in one go. 
            
            if Fuel == 'SOLAR PV' or Fuel == 'SOLAR THERMAL':
                  
                  main(inp3,row)
                  
            
            
               
            
            if Fuel == 'GEOTHERMAL BT' or Fuel == 'GEOTHERMAL FT':
                  
                  main(inp4,row)

            main(inp2,row)
            print(index)
            main(inp1,row)
            
            
            #THis block is for the input flows. 
            
            
            
            
            
            
            
            
            
            
#directory = 'C:/Users/tghosh/Desktop/LCI work/Day 63/New Folder/'
            
            
            
            
wb.save(filename)
            
            
            
print('Compilation Successful')       
            
#filename2= 'globalsave.pkl'

#dill.dump_session(filename2)



