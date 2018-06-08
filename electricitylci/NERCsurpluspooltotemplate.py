#Generates surplus pools process for NERC regions

import pandas as pd
from copy import copy
import openpyxl

#import globals
#datadir = globals.datadir
#outputdir = globals.outputdir

egrid = pd.read_csv(datadir+"egrid.csv", header=0, error_bad_lines=False)
temp = egrid[['NERC region acronym',' eGRID subregion acronym ']]
temp = temp.drop_duplicates()

primemover = pd.read_csv(datadir+"prime_moverlist.csv", header=0, error_bad_lines=False)
fuellist = pd.read_csv(datadir+"fuel_list.csv", header=0, error_bad_lines=False)
fuel_name = pd.read_csv(datadir+"fuelname.csv", header=0, error_bad_lines=False)
fuellist = fuellist.dropna(axis = 0, subset = ['FuelList'])
fuellist = fuellist.dropna(axis = 0, subset = [' eGRID subregion acronym '])





#THis is for inputting the region name from the user for which the templates will be built. 

Reg = input('Please enter 4 lettered NERC region here in uppercase: ')    

print('Let the magic begin ---- ')
#Read the template files. 
                
wb = openpyxl.load_workbook(datadir+'template.xlsx')
gi = wb['General information']
io = wb['InputsOutputs']

wb2 = openpyxl.load_workbook(datadir+'ConsumptionMix.xlsx')
data = wb2['Sheet1']

reg1 = data['A4':'A29']
trade = data['F4':'F29']

matrix = data['G3':'U29']





global blank_row
blank_row = 6
Column_number = 43 #THis is based on the template number of columns in the template.
   

index = 2; #Starting the index at a specific point
#REading the list of fuel for the entered region


               

#This block is used for filling up the General Infomration Sheet


        

cn = 3;

gi['D4'].value = 'Electricity '

gi['D5'].value = 'from Surplus Pool Mix';

gi['D6'].value = 'at region '+ Reg;

gi['D7'].value = 'Consumption Mix';

gi['D10'].value = 'Electricity from Surplus in the '+Reg+' region'

#This function is used for the name of the fuels.
def pm1(fn):
    for row in primemover.itertuples():
       if row[3]  == fn or row[4] == fn or row[5] == fn or row[6] == fn:
          if row[4] != None:                   
             return row[4].capitalize()

gi['D11'].value = '22: Utilities/2211: Electric Power Generation, Transmission and Distribution/'

gi['D12'].value = 'FALSE'

gi['D14'].value = 'Electricity; voltage?'

gi['D16'].value = '01/01/2017'

gi['D17'].value = '12/31/2017'

gi['D18'].value = '2014'

gi['D20'].value = 'US-eGRID-'+Reg

v= '' ;
#THis function is used for finding the NERC regions. THese input files are built manually. 
for row in temp.itertuples():
    if row[2] == Reg:
       v = row[1]

gi['D24'].value = 'EGRID subregion definitions:<https://www.epa.gov/energy/emissions-generation-resource-integrated-database-egrid>\nNERC region: '+v+'\nStates totally or partially included: <autofill>'

#This is for printing the Prime Movers. Written to a output file and reread back. 
def pm(fn,rg):
    f = open('out.txt', 'w')
    for row in primemover.itertuples():
      if row[2] == rg:
        if row[3]  == fn or row[4] == fn or row[5] == fn or row[6] == fn:
          if row[7] != None:                   
             f.write(row[7])
             f.write(',')
    f.close()        

 


#linestring = open('out.txt', 'r').read()

gi['D26'].value = 'This is an aggregation of technology types within this EGRID subregion.'

filename = outputdir + 'Surplus '+ Reg + '.xlsx'



#This block is used for filling up the InputOutput Sheet ONLY EMISSIONS


#assuming the starting row to fill in the emissions is the 5th row Blank

blank_row = 6;

#Reading All the files
io['D5'].value =  Reg+' surplus pool electricity';   

#This function is necessary to creates new line in the template for writing input and output flows. 
def createblnkrow(br):
    for i in range(1,Column_number):
      
      v = io.cell(row = br,column = i).value
      io.cell(row = br+1 ,column = i).value = v
      io.cell(row = br+1 ,column = i).font = copy(io.cell(row = br,column = i).font)
      io.cell(row = br+1 ,column = i).border = copy(io.cell(row = br,column = i).border)
      io.cell(row = br+1 ,column = i).fill = copy(io.cell(row = br,column = i).fill)
    br = br + 1;
      #print(io.cell(row = 6,column = 4).value);
    return br

row = blank_row;
col = 0;
 
index = 2;
            

for i in range(0,26):
   if Reg == reg1[i][0].value:
      if trade[i][0].value != 0:
           for j in range(0,7):
               
               if matrix[i+1][j].value != None and matrix[i+1][j].value !=0:
                           #print(matrix[i+1][j].value)
                           blank_row = createblnkrow(blank_row)
                           io[row][0].value = index;
                           io[row][2].value = 5;
                           io[row][3].value = 'Canada Provinces-'+matrix[0][j].value+'electricity';
                           io[row][4].value = '22: Utilities/2211: Electric Power Generation, Transmission and Distribution/'
                           io[row][6].value =  matrix[i+1][j].value
                           io[row][7].value = 'MWh';
                           io[row][21].value = 'Electricity Trade';                           
                           row = row+1;
                           index = index+1;
                           
           for j in range(7,8):
               
                if matrix[i+1][j].value != None and matrix[i+1][j].value !=0:
                           #print(matrix[i+1][j].value)
                           blank_row = createblnkrow(blank_row)
                           io[row][0].value = index;
                           io[row][2].value = 5;
                           io[row][3].value = 'Mexico-'+matrix[0][j].value+'electricity';
                           io[row][4].value = '22: Utilities/2211: Electric Power Generation, Transmission and Distribution/'
                           io[row][6].value =  matrix[i+1][j].value
                           io[row][7].value = 'MWh';
                           io[row][21].value = 'Electricity Trade';                           
                           row = row+1;
                           index = index+1; 
                           
           for j in range(8,15):
               
                if matrix[i+1][j].value != None and matrix[i+1][j].value !=0:
                           #print(matrix[i+1][j].value)
                           blank_row = createblnkrow(blank_row)
                           io[row][0].value = index;
                           io[row][2].value = 5;
                           io[row][3].value = 'Electricity 100 - 120V - '+matrix[0][j].value;
                           io[row][4].value = '22: Utilities/2211: Electric Power Generation, Transmission and Distribution/'
                           io[row][6].value =  matrix[i+1][j].value
                           io[row][7].value = 'MWh';
                           io[row][21].value = 'Electricity Trade';                           
                           row = row+1;
                           index = index+1;
#This function  is to find all other output flows and input fuels.


            
            
wb.save(filename)
            

            
print('Compilation Successful')       
            
#filename2= 'globalsave.pkl'

#dill.dump_session(filename2)



