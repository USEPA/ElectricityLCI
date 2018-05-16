#Writes out generation processes. Not currently working because relies on inventory from old sources

import pandas as pd
from copy import copy, deepcopy
import openpyxl
import numpy as np

def gen1(a,b):

        #Reading the facility file
        egrid1 = pd.read_csv("egrid_2014_1.csv", header=0, error_bad_lines=False)
        
        
        #Reading the flow by facility file
        egrid2 = pd.read_csv("egrid_2014.csv", header=0, error_bad_lines=False)
        
        
        egrid3 = egrid2.pivot(index = 'FacilityID',columns = 'FlowName', values = 'FlowAmount').reset_index()
        
        
        #Merging dataframes together
        egrid = egrid1.merge(egrid3, left_on = 'FacilityID', right_on = 'FacilityID')
        
        
        
        
        '''
        reg = egrid[['eGRID subregion acronym']]
        fuellist = egrid[["Plant primary fuel"]]
        fuelname = egrid[["Plant primary coal/oil/gas/ other fossil fuel category"]]
        state = egrid[["State"]]
        heatinput = egrid[['Heat input']]
        netgen = egrid[['Net generation']]
        '''
        egrid[['Heat input','Net generation']] = egrid[['Heat input','Net generation']].apply(pd.to_numeric,errors = 'coerce')
        
        
        egrid['Carbon dioxide'] = egrid['Carbon dioxide']/egrid['Net generation']
        egrid['Sulfur dioxide'] = egrid['Sulfur dioxide']/egrid['Net generation']
        egrid['Methane'] = egrid['Methane']/egrid['Net generation']
        egrid['Nitrogen oxides'] = egrid['Nitrogen oxides']/egrid['Net generation']
        egrid['Nitrous oxide'] =  egrid['Nitrous oxide']/egrid['Net generation']
        
        
        
        
        #Calculating efficiency
        egrid.loc[:,"Efficiency"] = egrid['Net generation'].values*100/egrid['Heat input'].values
        egrid = egrid.sort_values(by = ['Efficiency'], axis=0, ascending=True, inplace=False, kind='quicksort', na_position='last')
        #Replace inf
        egrid = egrid.replace([np.inf, -np.inf], np.nan)
        egrid = egrid.dropna(subset = ['Efficiency'])
        #Dropping less than 10%
        egrid_new1 = egrid[(egrid['Efficiency'] >= a) & (egrid['Efficiency'] <= b)]
        #egrid_new1.loc[:,'Nitrogen oxides','Carbon dioxide','Methane','Sulfur dioxide','Nitrous oxide'] = egrid_new1['Nitrogen oxides','Carbon dioxide','Methane','Sulfur dioxide','Nitrous oxide'].values/egrid['Net generation'].values
        #egrid_new1.loc[:,'Nitrogen oxides'] = egrid_new1['Nitrogen oxides'].values/egrid['Net generation'].values
      
        egrid_new = egrid_new1.drop(['Heat input','Net generation','Efficiency'],axis = 1)
        
        return egrid_new




#This function is necessary to creates new line in the template for writing input and output flows. 
def createblnkrow(br,io):
    
    Column_number = 43 #THis is based on the template number of columns in the template. 
    for i in range(1,Column_number):
      
      v = io.cell(row = br,column = i).value
      io.cell(row = br+1 ,column = i).value = v
      io.cell(row = br+1 ,column = i).font = copy(io.cell(row = br,column = i).font)
      io.cell(row = br+1 ,column = i).border = copy(io.cell(row = br,column = i).border)
      io.cell(row = br+1 ,column = i).fill = copy(io.cell(row = br,column = i).fill)
    br = br + 1;
      #print(io.cell(row = 6,column = 4).value);
    return br


'''
#This function is to find standard deviation values. 
def standdev(d1,em):
    for r in d1.itertuples():
         if r[3] == Fuel and em == r[4]: 
           if r[2] == Reg and pd.isnull(r[5]) == False and r[5] != 0:
                return r[5];
'''






#THis is for inputting the region name from the user for which the templates will be built. 





#REading the list of fuel for the entered region


def generator(a,b):
    
            egrid = gen1(a,b)
            Reg = 'AZNM'
            #Reg = input('Please enter 4 lettered region here in uppercase: ')  
            egrid = egrid[egrid['eGRID subregion acronym'] == Reg]
            
            fuel_name = pd.read_csv("fuelname.csv", header=0, error_bad_lines=False)
            for row in fuel_name.itertuples():
              #assuming the starting row to fill in the emissions is the 5th row Blank    
              global blank_row;    
              blank_row = 6;
              index = 2;
              for roww in egrid.itertuples():
                
                if row[1] == roww[8] or row[1] == roww[7]:
                    
                    
                    fuelname = row[2]
                    print(fuelname)
                    
                    
        
                    #Read the template files.                         
                    wbook = openpyxl.load_workbook('Fed_template_Main_19Dec2016.xlsx')
                    gi = wbook['General information']
                    
                    #This block is used for filling up the General Infomration Sheet
                    
                    
                            
                    
                    
                    
                    gi['D4'].value = 'Electricity'
                    
                    gi['D5'].value = 'from '+ fuelname;
                    
                    gi['D6'].value = 'at generating facility';
                    
                    gi['D7'].value = 'Production Mix'; 
                    
                    
                    
                    gi['D10'].value = 'Electricity from '+ fuelname+' using power plants in the '+Reg+' region'
                    
                    #This function is used for the name of the fuels.
                    #def pm1(fn):
                    #    for row in primemover.itertuples():
                    #       if row[3]  == fn or row[4] == fn or row[5] == fn or row[6] == fn:
                    #          if row[4] != None:                   
                    #             return row[4].capitalize()
                    
                    gi['D11'].value = '22: Utilities/2211: Electric Power Generation, Transmission and Distribution/'+str(fuelname)
                    
                    gi['D12'].value = 'FALSE'
                    
                    gi['D14'].value = 'Electricity; voltage?'
                    
                    gi['D16'].value = '01/01/2017'
                    
                    gi['D17'].value = '12/31/2017'
                    
                    gi['D18'].value = '2014'
                    
                    gi['D20'].value = 'US-eGRID-'+Reg
                
                    v= '' ;
                    v1= [];
                    '''
                    #THis function is used for finding the NERC regions. THese input files are built manually. 
                    for row in temp.itertuples():
                        if row[2] == Reg:
                           v = row[1]
                           
                    for row in state.itertuples():
                        if row[1] == Reg:
                            v1 = row[2];
                    
                    gi['D24'].value = 'EGRID subregion definitions:<https://www.epa.gov/energy/emissions-generation-resource-integrated-database-egrid>\nNERC region: '+v+'\nStates totally or partially included: '+v1
                    '''
                    for rowww in egrid.itertuples():
                        if row[1] == roww[8]:
                            v1.append(rowww[2])
                            
                    v2 = list(set(v1))        
                    str1 = ','.join(v2)
                    
                    gi['D24'].value = 'EGRID subregion definitions:<https://www.epa.gov/energy/emissions-generation-resource-integrated-database-egrid>\nNERC region: '+v+'\nStates totally or partially included: '+str1
                   
                    
                    '''
                    #This is for printing the Prime Movers. Written to a output file and reread back. 
                    def pm(fn,rg):
                        f = open('out.txt', 'w')
                        for row in primemover.itertuples():
                          if row[2] == rg:
                            if row[3]  == fn or row[4] == fn or row[5] == fn or row[6] == fn:
                              if row[7] != None:                   
                                 f.write(row[7])
                                 f.write(',')
                        f.close()        
                    
                     
                    pm(Fuel,Reg)
                    
                    #linestring = open('out.txt', 'r').read()
                    
                    gi['D26'].value = 'This is an aggregation of technology types within this EGRID subregion.  List of prime movers:'+linestring
                    '''
                    
                    
               

                    egrid_f1 = egrid[egrid['Plant primary coal/oil/gas/ other fossil fuel category'] == row[1]]
                    if egrid_f1.empty == True:
                          egrid_f1 = egrid[egrid['Plant primary fuel'] == row[1]]              
                    
    
                    #This block is used for filling up the InputOutput Sheet ONLY EMISSIONS
                    io = wbook['InputsOutputs']
                    


                    #Fillin up with information emission only
                    io['D5'].value =  'Electricity from '+fuelname;  
                    io['G5'].value =  1 
                    io['H5'].value =  'MJ'
                    row1 = blank_row;
                


                    flownames = list(egrid_f1.columns.values)
                    
                    for i in range(8,len(flownames)):
                    
                        if np.isnan(egrid_f1[flownames[i]].mean()) == False:
                            
                            blank_row = createblnkrow(blank_row,io)
                            io[row1][0].value = index;
                            io[row1][2].value = 4;
                            io[row1][3].value = flownames[i];
                            io[row1][4].value = 'Elementary Flows/air/unspecified'
                            io[row1][6].value = egrid_f1[flownames[i]].mean();
                            io[row1][7].value = 'kg';
                            io[row1][21].value = 'EGRID data with plants over 10% efficiency';
                            io[row1][26].value = 'Normal';
                            io[row1][27].value = egrid_f1[flownames[i]].mean();
                            io[row1][28].value = egrid_f1[flownames[i]].std();
                            row1 = row1+1;
                            index = index+1;

                    filename = fuelname+"_"+Reg+".xlsx"
                    wbook.save(filename)
                    break;
generator(10,100)

'''
            
            

#This block is used for filling up the InputOutput Sheet ONLY EMISSIONS
io = wb.get_sheet_by_name('InputsOutputs')

#assuming the starting row to fill in the emissions is the 5th row Blank

global blank_row;

blank_row = 6;
Column_number = 43 #THis is based on the template number of columns in the template. 

#Reading All the files
io['D5'].value =  'Electricity from '+fuelname;   
lci1 = pd.read_csv("FinalLCIforimport1.csv", header=0, error_bad_lines=False)
lci2 = pd.read_csv("FinalLCIforimport2.csv", header=0, error_bad_lines=False)
lci3 = pd.read_csv("FinalLCIforimport3.csv", header=0, error_bad_lines=False)
lci4 = pd.read_csv("FinalLCIforimport4.csv", header=0, error_bad_lines=False)
nei1 = pd.read_csv("FinalNEIforimport1.csv", header=0, error_bad_lines=False)
nei2 = pd.read_csv("FinalNEIforimport2.csv", header=0, error_bad_lines=False)
nei3 = pd.read_csv("FinalNEIforimport3.csv", header=0, error_bad_lines=False)
nei4 = pd.read_csv("FinalNEIforimport4.csv", header=0, error_bad_lines=False)
Atri1 = pd.read_csv("Finaltriforimport1air.csv", header=0, error_bad_lines=False)
Atri2 = pd.read_csv("Finaltriforimport2air.csv", header=0, error_bad_lines=False)
Atri3 = pd.read_csv("Finaltriforimport3air.csv", header=0, error_bad_lines=False)
Atri4 = pd.read_csv("Finaltriforimport4air.csv", header=0, error_bad_lines=False)
Wtri1 = pd.read_csv("Finaltriforimport1water.csv", header=0, error_bad_lines=False)
Wtri2 = pd.read_csv("Finaltriforimport2water.csv", header=0, error_bad_lines=False)
Wtri3 = pd.read_csv("Finaltriforimport3water.csv", header=0, error_bad_lines=False)
Wtri4 = pd.read_csv("Finaltriforimport4water.csv", header=0, error_bad_lines=False)
Stri1 = pd.read_csv("Finaltriforimport1soil.csv", header=0, error_bad_lines=False)
Stri2 = pd.read_csv("Finaltriforimport2soil.csv", header=0, error_bad_lines=False)
Stri3 = pd.read_csv("Finaltriforimport3soil.csv", header=0, error_bad_lines=False)
Stri4 = pd.read_csv("Finaltriforimport4soil.csv", header=0, error_bad_lines=False)

sdlci1 = pd.read_csv("FinalLCIsdforimport1.csv", header=0, error_bad_lines=False)
sdlci2 = pd.read_csv("FinalLCIsdforimport2.csv", header=0, error_bad_lines=False)
sdlci3 = pd.read_csv("FinalLCIsdforimport3.csv", header=0, error_bad_lines=False)
sdlci4 = pd.read_csv("FinalLCIsdforimport4.csv", header=0, error_bad_lines=False)
sdnei1 = pd.read_csv("FinalNEIsdforimport1.csv", header=0, error_bad_lines=False)
sdnei2 = pd.read_csv("FinalNEIsdforimport2.csv", header=0, error_bad_lines=False)
sdnei3 = pd.read_csv("FinalNEIsdforimport3.csv", header=0, error_bad_lines=False)
sdnei4 = pd.read_csv("FinalNEIsdforimport4.csv", header=0, error_bad_lines=False)
sdAtri1 = pd.read_csv("Finalsdtriforimportair1.csv", header=0, error_bad_lines=False)
sdAtri2 = pd.read_csv("Finalsdtriforimportair2.csv", header=0, error_bad_lines=False)
sdAtri3 = pd.read_csv("Finalsdtriforimportair3.csv", header=0, error_bad_lines=False)
sdAtri4 = pd.read_csv("Finalsdtriforimportair4.csv", header=0, error_bad_lines=False)
sdWtri1 = pd.read_csv("Finalsdtriforimportwater1.csv", header=0, error_bad_lines=False)
sdWtri2 = pd.read_csv("Finalsdtriforimportwater2.csv", header=0, error_bad_lines=False)
sdWtri3 = pd.read_csv("Finalsdtriforimportwater3.csv", header=0, error_bad_lines=False)
sdWtri4 = pd.read_csv("Finalsdtriforimportwater4.csv", header=0, error_bad_lines=False)
sdStri1 = pd.read_csv("Finalsdtriforimportsoil1.csv", header=0, error_bad_lines=False)
sdStri2 = pd.read_csv("Finalsdtriforimportsoil2.csv", header=0, error_bad_lines=False)
sdStri3 = pd.read_csv("Finalsdtriforimportsoil3.csv", header=0, error_bad_lines=False)
sdStri4 = pd.read_csv("Finalsdtriforimportsoil4.csv", header=0, error_bad_lines=False)
fuelinp1 = pd.read_csv("AggregatedFuelinput1.csv", header=0, error_bad_lines=False)
fuelinp2 = pd.read_csv("AggregatedFuelinput2.csv", header=0, error_bad_lines=False)
fuelinp3 = pd.read_csv("AggregatedFuelinput3.csv", header=0, error_bad_lines=False)
fuelinp4 = pd.read_csv("AggregatedFuelinput4.csv", header=0, error_bad_lines=False)

#This function is necessary to creates new line in the template for writing input and output flows. 
def createblnkrow(br):
    for i in range(1,Column_number):
      
      v = io.cell(row = br,column = i).value
      io.cell(row = br+1 ,column = i).value = v
      io.cell(row = br+1 ,column = i).font = copy(io.cell(row = br,column = i).font)
      io.cell(row = br+1 ,column = i).border = copy(io.cell(row = br,column = i).border)
      io.cell(row = br+1 ,column = i).fill = copy(io.cell(row = br,column = i).fill)
    br = br + 1;
      #print(io.cell(row = 6,column = 4).value);
    return br

row = blank_row;
col = 0;

index = 2; #Starting the index at a specific point

#This function is to find standard deviation values. 
def standdev(d1,em):
    for r in d1.itertuples():
         if r[3] == Fuel and em == r[4]: 
           if r[2] == Reg and pd.isnull(r[5]) == False and r[5] != 0:
                return r[5];
            




#This function  is to find all other output flows and input fuels.
def main(d1,d2,d3,d4,d5,d6,d7,d8,d9,d10,d11,row,index):
     global blank_row;
     for r in d1.itertuples():
         if r[3] == Fuel: 
           if r[2] == Reg and pd.isnull(r[4]) == False:
               blank_row = createblnkrow(blank_row)
               io[row][0].value = index;
               io[row][2].value = 5;
               io[row][3].value = fuelname;
               io[row][4].value = 'Category of the fuel flow'
               io[row][6].value = r[4];
               io[row][7].value = 'mmBTU';
               io[row][21].value = 'Fuel Input';
               
               row = row+1;
               index = index+1; 

     for r in d2.itertuples():
         if r[3] == Fuel: 
           if r[2] == Reg and pd.isnull(r[5]) == False and r[5] != 0:
               blank_row = createblnkrow(blank_row)
               io[row][0].value = index;
               io[row][2].value = 4;
               io[row][3].value = r[4];
               io[row][4].value = 'Elementary Flows/air/unspecified'
               io[row][6].value = r[5];
               io[row][7].value = 'kg';
               io[row][21].value = 'EGRID data with plants over 10% efficiency';
               io[row][26].value = 'Normal';
               io[row][27].value = r[5];
               io[row][28].value = standdev(d7,r[4])
               row = row+1;
               index = index+1;
               
     for r in d3.itertuples():
         if r[3] == Fuel: 
           if r[2] == Reg and pd.isnull(r[5]) == False and r[5] != 0:
               blank_row = createblnkrow(blank_row)
               io[row][0].value = index;
               io[row][2].value = 4;
               io[row][3].value = r[4];
               io[row][4].value = 'Elementary Flows/air/unspecified'
               io[row][6].value = r[5];
               io[row][7].value = 'kg';
               io[row][21].value = 'NEI data with plants over 10% efficiency';
               io[row][26].value = 'Normal';
               io[row][27].value = r[5];
               io[row][28].value = standdev(d8,r[4])
               row = row+1;
               index = index+1;         
  
     for r in d4.itertuples():
         if r[3] == Fuel: 
           if r[2] == Reg and pd.isnull(r[5]) == False and r[5] != 0:
               blank_row = createblnkrow(blank_row)
               io[row][0].value = index;
               io[row][2].value = 4;
               io[row][3].value = r[4];
               io[row][4].value = 'Elementary Flows/air/unspecified'
               io[row][6].value = r[5];
               io[row][7].value = 'kg';
               io[row][21].value = 'TRI data with plants over 10% efficiency';
               io[row][26].value = 'Normal';
               io[row][27].value = r[5];
               io[row][28].value = standdev(d9,r[4])
               row = row+1;
               index = index+1;                


     for r in d5.itertuples():
         if r[3] == Fuel: 
           if r[2] == Reg and pd.isnull(r[5]) == False and r[5] != 0:
               blank_row = createblnkrow(blank_row)
               io[row][0].value = index;
               io[row][2].value = 4;
               io[row][3].value = r[4];
               io[row][4].value = 'Elementary Flows/water/unspecified'
               io[row][6].value = r[5];
               io[row][7].value = 'kg';
               io[row][21].value = 'TRI data with plants over 10% efficiency';
               io[row][26].value = 'Normal';
               io[row][27].value = r[5];
               io[row][28].value = standdev(d10,r[4])
               row = row+1;
               index = index+1; 

     for r in d6.itertuples():
         if r[3] == Fuel: 
           if r[2] == Reg and pd.isnull(r[5]) == False and r[5] != 0:
               blank_row = createblnkrow(blank_row)
               io[row][0].value = index;
               io[row][2].value = 4;
               io[row][3].value = r[4];
               io[row][4].value = 'Elementary Flows/soil/unspecified'
               io[row][6].value = r[5];
               io[row][7].value = 'kg';
               io[row][21].value = 'TRI data with plants over 10% efficiency';
               io[row][26].value = 'Normal';
               io[row][27].value = r[5];
               io[row][28].value = standdev(d11,r[4])
               row = row+1;
               index = index+1;


#THis block is used for inputting input flows                        
inputs = pd.read_csv("importinputdata.csv", header=0, error_bad_lines=False)

for r in inputs.itertuples():
         if r[2] == Fuel and r[4] != None: 
               blank_row = createblnkrow(blank_row)
               io[row][0].value = index;
               io[row][2].value = 5;
               io[row][3].value = r[3];
               io[row][4].value = 'Category of Input FLOWS'
               io[row][6].value = r[4];
               if r[2] == 'Power Plant [Construction]':
                   io[row][7].value = 'pce';
               else:
                   io[row][7].value = 'kg';
               io[row][21].value = 'NEI DATA';
               row = row+1;
               index = index+1;
   


#Separate calls are made for Solar and Geothermal Fuel sources because of different databases. 
#Different databases are sent because many information have to be written in one go. 

if Fuel == 'SOLAR PV' or Fuel == 'SOLAR THERMAL':
      
      main(fuelinp3,lci3,nei3,Atri3,Wtri3,Stri3,sdlci3,sdnei3,sdAtri3,sdWtri3,sdStri3,row,index)
      


   

if Fuel == 'GEOTHERMAL BT' or Fuel == 'GEOTHERMAL FT':
      
      main(fuelinp4,lci4,nei4,Atri4,Wtri4,Stri4,sdlci4,sdnei4,sdAtri4,sdWtri4,sdStri4,row,index)
      

main(fuelinp2,lci2,nei2,Atri2,Wtri2,Stri2,sdlci2,sdnei2,sdAtri2,sdWtri2,sdStri2,row,index)

main(fuelinp1,lci1,nei1,Atri1,Wtri1,Stri1,sdlci1,sdnei1,sdAtri1,sdWtri1,sdStri1,row,index)


#THis block is for the input flows. 










directory = 'C:/Users/tghosh/Desktop/LCI work/Day 62/TJ internship work/8. Database Generation/Final/files/'




wb.save(directory+filename)



print('Compilation Successful')       

#filename2= 'globalsave.pkl'

#dill.dump_session(filename2)

'''

