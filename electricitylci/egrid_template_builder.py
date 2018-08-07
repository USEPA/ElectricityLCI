from electricitylci.egrid_facilities import egrid_subregions
from electricitylci.globals import fuel_name
import os
import openpyxl
from copy import copy, deepcopy


#This function is necessary to creates new line in the template for writing input and output flows. 
def createblnkrow(br,io):
    
    Column_number = 43 #THis is based on the template number of columns in the template. 
    for i in range(1,Column_number):
      
      v = io.cell(row = br,column = i).value
      io.cell(row = br+1 ,column = i).value = v
      io.cell(row = br+1 ,column = i).font = copy(io.cell(row = br,column = i).font)
      io.cell(row = br+1 ,column = i).border = copy(io.cell(row = br,column = i).border)
      io.cell(row = br+1 ,column = i).fill = copy(io.cell(row = br,column = i).fill)
    
    br = br + 1;
      #print(io.cell(row = 6,column = 4).value);
    return br




def gen_process_template_generator(generation_process_dict):
        
        data_dir = os.path.dirname(os.path.realpath(__file__))   
        os.chdir(data_dir+'/data/')
        for Reg in egrid_subregions: 
         
            
          for index,row in fuel_name.iterrows(): 
        #assuming the starting row to fill in the emissions is the 5th row Blank
           
           #Reading complete fuel name and heat content information       
           fuelname = row['Fuelname'] 
             
           if (Reg+'_'+fuelname) in generation_process_dict: 
            global blank_row;    
            blank_row = 6;
            
            row1 = 5;
            
             
      
            #Opening tempalte file for wriitng
            wbook = openpyxl.load_workbook('template.xlsx')
            #Read the template files.                         
        
            gi = wbook['General information']          
            
            
            #This block is used for filling up the General Infomration Sheet      
            gi['D4'].value = 'Electricity'
            
            gi['D5'].value = 'from '+ fuelname;
            
            gi['D6'].value = 'at generating facility';
            
            gi['D10'].value = 'Electricity from '+ fuelname+' using power plants in the '+Reg+' region'
            
            #This function is used for the name of the fuels.
            #def pm1(fn):
            #    for row in primemover.itertuples():
            #       if row[3]  == fn or row[4] == fn or row[5] == fn or row[6] == fn:
            #          if row[4] != None:                   
            #             return row[4].capitalize()
            
            gi['D11'].value = generation_process_dict[Reg+'_'+fuelname]['category']
            
            gi['D12'].value = 'FALSE'
            
            gi['D14'].value = 'Electricity; voltage?'
            
            gi['D16'].value = generation_process_dict[Reg+'_'+fuelname]['processDocumentation']['validFrom']
            
            gi['D17'].value = generation_process_dict[Reg+'_'+fuelname]['processDocumentation']['validUntil']
            
                
            gi['D18'].value = generation_process_dict[Reg+'_'+fuelname]['processDocumentation']['validFrom']
            
            gi['D20'].value = 'US-eGRID-'+Reg
        
            gi['D21'].value = 'F'
            v= [];
            v1= [];
            
            
          
            
            gi['D24'].value = 'database subregion definitions:<https://www.epa.gov/energy/emissions-generation-resource-integrated-database-database>\nNERC region: '
                                  
            gi['D26'].value = 'This is an aggregation of technology types for this fuel type within this eGRID subregion'
            
            
        
        
            
        
            io = wbook['InputsOutputs']
        
            
            for index in range(0,len(generation_process_dict[Reg+'_'+fuelname]['exchanges'])):
                
                #print(generation_process_dict[Reg+'_'+fuelname]['exchanges'][index])
            
        
                blank_row = createblnkrow(blank_row,io)
                
                
                io[row1][0].value = index+1;
                if generation_process_dict[Reg+'_'+fuelname]['exchanges'][index]['input'] == True:
                  io[row1][1].value = 5
                else: 
                  if  index == 0: 
                     io[row1][2].value = 0
                  else:
                     io[row1][2].value = 4
                
                if index == 0:
                   name = generation_process_dict[Reg+'_'+fuelname]['exchanges'][index]['flow']['name']+' from '+str(fuelname)
                else:
                   name = generation_process_dict[Reg+'_'+fuelname]['exchanges'][index]['flow']['name']
                    
                if generation_process_dict[Reg+'_'+fuelname]['exchanges'][index]['flow']['name'] == 'Steam':
                   io[row1][2].value = 2 
                #Making the string flow name within limits accepted by OpenLCA. 
                io[row1][3].value  = name[0:255]
                
                if index == 0:
                  io[row1][4].value = generation_process_dict[Reg+'_'+fuelname]['category']
                  io[row1][5].value = 'US-eGRID-'+generation_process_dict[Reg+'_'+fuelname]['location']['name']
                else:
                  io[row1][4].value = generation_process_dict[Reg+'_'+fuelname]['exchanges'][index]['category']
                  if generation_process_dict[Reg+'_'+fuelname]['exchanges'][index]['category'] == 'Waste flows/':
                    io[row1][2].value = 3  
                    
               
                io[row1][6].value = generation_process_dict[Reg+'_'+fuelname]['exchanges'][index]['amount']
                io[row1][7].value = generation_process_dict[Reg+'_'+fuelname]['exchanges'][index]['unit']['name']
                
                if index > 1:
                    
                    io[row1][12].value = generation_process_dict[Reg+'_'+fuelname]['exchanges'][index]['dqEntry'][1:4]
                    io[row1][13].value = generation_process_dict[Reg+'_'+fuelname]['exchanges'][index]['dqEntry'][5]
                    io[row1][14].value = generation_process_dict[Reg+'_'+fuelname]['exchanges'][index]['dqEntry'][7]
                    io[row1][15].value = generation_process_dict[Reg+'_'+fuelname]['exchanges'][index]['dqEntry'][9]
                io[row1][21].value = 'database data with plants over 10% efficiency';
                
                if 'uncertainty' in generation_process_dict[Reg+'_'+fuelname]['exchanges'][index]:
                   io[row1][26].value = generation_process_dict[Reg+'_'+fuelname]['exchanges'][index]['uncertainty']['distributionType']
                   if 'geomMean' in generation_process_dict[Reg+'_'+fuelname]['exchanges'][index]['uncertainty'] and 'geomSd' in generation_process_dict[Reg+'_'+fuelname]['exchanges'][index]['uncertainty']: 
                    #uncertianty calculations
                    io[row1][27].value = generation_process_dict[Reg+'_'+fuelname]['exchanges'][index]['uncertainty']['geomMean']
                    io[row1][28].value = generation_process_dict[Reg+'_'+fuelname]['exchanges'][index]['uncertainty']['geomSd']
                   io[row1][29].value = generation_process_dict[Reg+'_'+fuelname]['exchanges'][index]['uncertainty']['maximum']
                   io[row1][30].value = generation_process_dict[Reg+'_'+fuelname]['exchanges'][index]['uncertainty']['minimum']
                if 'comment' in generation_process_dict[Reg+'_'+fuelname]['exchanges'][index]:
                    io[row1][33].value = generation_process_dict[Reg+'_'+fuelname]['exchanges'][index]['comment']
        
                
                row1 = blank_row-1
                
                
                    
        
        
                     
                
                 
                    
            #Writing final file. 
            
            filename = fuelname+"_"+Reg+".xlsx"
            data_dir = os.path.dirname(os.path.realpath(__file__))+"/output/"
            wbook.save(data_dir+filename)
            print(filename+' File written Successfully')
 

def gen_mix_template_generator(generation_mix_dict):
        
        data_dir = os.path.dirname(os.path.realpath(__file__))   
        os.chdir(data_dir+'/data/')
        

        
        for Reg in egrid_subregions:       
            
                #Opening tempalte file for writing
                wbook = openpyxl.load_workbook('template.xlsx')           
                         
                global blank_row 
                blank_row = 6
                
                row1 = 5
                
                 
       
             
                #Read the template files.                         
            
                gi = wbook['General information']          
                
                
                #This block is used for filling up the General Infomration Sheet      
                    #This block is used for filling up the General Infomration Sheet      
                gi['D4'].value = 'Electricity'
            
                gi['D5'].value = 'from Generation';
            
                gi['D6'].value = 'at region '+Reg;
            
            #Mix type
                gi['D7'].value = 'Production Mix'; 
            
                gi['D9'].value = 'Electricity; from Generation mix; at region '+Reg; 
            
                gi['D10'].value = 'Electricity from generation mix using power plants in the '+Reg+' region'
            
            #This function is used for the name of the fuels.
            #def pm1(fn):
            #    for row in primemover.itertuples():
            #       if row[3]  == fn or row[4] == fn or row[5] == fn or row[6] == fn:
            #          if row[4] != None:                   
            #             return row[4].capitalize()
            
                gi['D11'].value = '22: Utilities/2211: Electric Power Generation, Transmission and Distribution/'
            
                gi['D12'].value = 'FALSE'
            
                gi['D14'].value = 'Electricity; voltage?'
                
                gi['D16'].value = generation_mix_dict[Reg]['processDocumentation']['validFrom']
                
                gi['D17'].value = generation_mix_dict[Reg]['processDocumentation']['validUntil']
                
                    
                gi['D18'].value = generation_mix_dict[Reg]['processDocumentation']['validFrom']
                
                gi['D20'].value = 'US-eGRID-'+Reg
            
                gi['D21'].value = 'F'
                v= [];
                v1= [];
                
                
              
                
                gi['D24'].value = 'database subregion definitions:<https://www.epa.gov/energy/emissions-generation-resource-integrated-database-database>\nNERC region: '
                                      
                gi['D26'].value = 'This is an aggregation of technology types within this eGRID subregion.  List of prime movers:'
                
                
            
            
                
            
                io = wbook['InputsOutputs']
            
                
                for index in range(0,len(generation_mix_dict[Reg]['exchanges'])):
                    
              
            
                    blank_row = createblnkrow(blank_row,io)
                    
                    
                    io[row1][0].value = index+1;
                    if generation_mix_dict[Reg]['exchanges'][index]['input'] == True:
                      io[row1][1].value = 5
                    else: 
                      if  index == 0: 
                         io[row1][2].value = 0
                      else:
                         io[row1][2].value = 5
                    
                    name = generation_mix_dict[Reg]['exchanges'][index]['flow']['name']
                    
                    #Making the string flow name within limits accepted by OpenLCA. 
                    io[row1][3].value  = name[0:255]
                    
                    if index == 0:
                      io[row1][3].value  = name[0:255]+' from Generation mix '+str(Reg)
                      io[row1][4].value = generation_mix_dict[Reg]['category']
                    else:
                      io[row1][4].value = generation_mix_dict[Reg]['exchanges'][index]['category']
                      
                    io[row1][5].value = 'US-eGRID-'+generation_mix_dict[Reg]['location']['name']
                    io[row1][6].value = generation_mix_dict[Reg]['exchanges'][index]['amount']
                    io[row1][7].value = generation_mix_dict[Reg]['exchanges'][index]['unit']['name']
                    io[row1][21].value = 'database data with plants over 10% efficiency';
                    
                    if 'comment' in generation_mix_dict[Reg]['exchanges'][index]:
                        io[row1][33].value = generation_mix_dict[Reg]['exchanges'][index]['comment']
            
                    
                    row1 = blank_row-1
                    
                    
                        
            
            
                         
                    
                     
                        
                    #Writing final file. 
                    
                filename = Reg+"_Generation_Mix.xlsx"
                data_dir = os.path.dirname(os.path.realpath(__file__))+"/output/"
                wbook.save(data_dir+filename)
                print(filename+' File written Successfully')




def distribution_template_generator(distribution_dict,efficiency):

        data_dir = os.path.dirname(os.path.realpath(__file__))   
        os.chdir(data_dir+'/data/')            
         
        
    
        for Reg in egrid_subregions:
                        
                            
                        wb = openpyxl.load_workbook('template.xlsx')
                        
                        gi = wb['General information']
                        io = wb['InputsOutputs']
                        
                        
                        global blank_row;
                                   
                        blank_row = 6;
                        Column_number = 43 #THis is based on the template number of columns in the template. 
                           
                        
                        index = 1; #Starting the index at a specific point
                        #REading the list of fuel for the entered region
                        
                        
                                       
                        
                        #This block is used for filling up the General Infomration Sheet
                        
                        
                                
                        
                        
                        
                        gi['D4'].value = 'Electricity '
                        
                        gi['D5'].value = 'from Distribution Mix';
                        
                        gi['D6'].value = 'at region '+ Reg;
                        
                        gi['D7'].value = 'Distribution Mix';
                        
                        gi['D10'].value = 'Distribution mix electricity in the '+Reg+' region'
                        
                        
                        gi['D11'].value = distribution_dict[Reg]['category']
                        
                        gi['D12'].value = 'FALSE'
                        
                        gi['D14'].value = 'Electricity; voltage?'
                        
                        gi['D16'].value = distribution_dict[Reg]['processDocumentation']['validFrom']
                        
                        gi['D17'].value = distribution_dict[Reg]['processDocumentation']['validUntil']
                        
                        gi['D18'].value = distribution_dict[Reg]['processDocumentation']['validFrom']
                        
                        gi['D20'].value = 'US-eGRID-'+Reg
                        
                        v= '' ;
                        #THis function is used for finding the NERC regions. THese input files are built manually. 
            
                        
                        gi['D24'].value = 'EGRID subregion definitions:<https://www.epa.gov/energy/emissions-generation-resource-integrated-database-egrid>\nNERC region: \nStates totally or partially included: <autofill>'
                        
                        #This is for printing the Prime Movers. Written to a output file and reread back. 
                        #def pm(fn,rg):
                        #    f = open('out.txt', 'w')
                        #    for row in primemover.itertuples():
                        ##      if row[2] == rg:
                         #       if row[3]  == fn or row[4] == fn or row[5] == fn or row[6] == fn:
                        #          if row[7] != None:                   
                        #             f.write(row[7])
                        #             f.write(',')
                        #    f.close()        
                        
                         
                        
                        
                        #linestring = open('out.txt', 'r').read()
                        
                        gi['D26'].value = 'This is the Distribution Mix.'
                        
                        filename = 'Distribution Mix '+ Reg + '.xlsx'
                        
                        
                        
                        #This block is used for filling up the InputOutput Sheet ONLY EMISSIONS
                        
                                               
                        
                        
                        #Reading All the files
                        #io['D5'].value =  Reg+' Distribution Mix electricity';   
                        
                        #This function is necessary to creates new line in the template for writing input and output flows. 
                        def createblnkrow(br):
                            for i in range(1,Column_number):
                              
                              v = io.cell(row = br,column = i).value
                              io.cell(row = br+1 ,column = i).value = v
                              io.cell(row = br+1 ,column = i).font = copy(io.cell(row = br,column = i).font)
                              io.cell(row = br+1 ,column = i).border = copy(io.cell(row = br,column = i).border)
                              io.cell(row = br+1 ,column = i).fill = copy(io.cell(row = br,column = i).fill)
                            br = br + 1;
                              #print(io.cell(row = 6,column = 4).value);
                            return br
                        
                        row = blank_row-1;
                        
                         
                        index = 1;
                        #Filling up the generation Mix flow at first                        
                        
                        io[row][0].value = index;
                        io[row][2].value = 0;
                        io[row][3].value = 'Electricity 100 - 120V';
                        io[row][4].value = distribution_dict[Reg]['category']
                        io[row][5].value = str(Reg)
                        io[row][6].value = 1
                        io[row][7].value = 'MWh';
                        io[row][21].value = 'Electricity Distribution Mix';                           
                        row = row+1;
                        index = index+1;
                        
                            

                        blank_row = createblnkrow(blank_row)
                        io[row][0].value = index;
                        io[row][1].value = 5;
                        io[row][3].value = distribution_dict[Reg]['exchanges'][index-1]['flow']['name']
                        io[row][4].value = distribution_dict[Reg]['category']
                        io[row][5].value = str(Reg)
                        io[row][6].value = 1/efficiency
                        io[row][7].value = 'MWh';
                        io[row][21].value = distribution_dict[Reg]['description']                         
                        row = row+1;
                        index = index+1;
                        
                                                     

                        
                                        
                             
                        filename = Reg+"_Distribution.xlsx"
                        data_dir = os.path.dirname(os.path.realpath(__file__))+"/output/"
                        wb.save(data_dir+filename)
                        print(filename+' File written Successfully')
                        
                



def consumption_mix_template_generator(consumption_dict):

        data_dir = os.path.dirname(os.path.realpath(__file__))   
        os.chdir(data_dir+'/data/')            
         
        
    
        for Reg in egrid_subregions:
                        
                            
                        wb = openpyxl.load_workbook('template.xlsx')
                        
                        gi = wb['General information']
                        io = wb['InputsOutputs']
                        
                        
                        global blank_row;
                                   
                        blank_row = 6;
                        Column_number = 43 #THis is based on the template number of columns in the template. 
                           
                        
                        index = 1; #Starting the index at a specific point
                        #REading the list of fuel for the entered region
                        
                        
                                       
                        
                        #This block is used for filling up the General Infomration Sheet
                        
                        
                                
                        
                        
                        
                        gi['D4'].value = 'Electricity '
                        
                        gi['D5'].value = 'for Consumption';
                        
                        gi['D6'].value = 'at region '+ Reg;
                        
                        gi['D7'].value = 'Consumption Mix';
                        
                        gi['D10'].value = 'Consumption mix electricity in the '+Reg+' region'
                        
                        
                        gi['D11'].value = consumption_dict[Reg]['category']
                        
                        gi['D12'].value = 'FALSE'
                        
                        gi['D14'].value = 'Electricity; voltage?'
                        
                        gi['D16'].value = consumption_dict[Reg]['processDocumentation']['validFrom']
                        
                        gi['D17'].value = consumption_dict[Reg]['processDocumentation']['validUntil']
                        
                        gi['D18'].value = consumption_dict[Reg]['processDocumentation']['validFrom']
                        
                        gi['D20'].value = 'US-eGRID-'+Reg
                        
                        v= '' ;
                        #THis function is used for finding the NERC regions. THese input files are built manually. 
            
                        
                        gi['D24'].value = 'EGRID subregion definitions:<https://www.epa.gov/energy/emissions-generation-resource-integrated-database-egrid>\nNERC region: \nStates totally or partially included: <autofill>'
                        
                        #This is for printing the Prime Movers. Written to a output file and reread back. 
                        #def pm(fn,rg):
                        #    f = open('out.txt', 'w')
                        #    for row in primemover.itertuples():
                        ##      if row[2] == rg:
                         #       if row[3]  == fn or row[4] == fn or row[5] == fn or row[6] == fn:
                        #          if row[7] != None:                   
                        #             f.write(row[7])
                        #             f.write(',')
                        #    f.close()        
                        
                         
                        
                        
                        #linestring = open('out.txt', 'r').read()
                        
                        gi['D26'].value = 'This is the Consumption Mix.'
                        
                        filename = 'Consumption Mix '+ Reg + '.xlsx'
                        
                        
                        
                        #This block is used for filling up the InputOutput Sheet ONLY EMISSIONS
                        
                                               
                        
                        
                        #Reading All the files
                        #io['D5'].value =  Reg+' Distribution Mix electricity';   
                        
                        #This function is necessary to creates new line in the template for writing input and output flows. 
                        def createblnkrow(br):
                            for i in range(1,Column_number):
                              
                              v = io.cell(row = br,column = i).value
                              io.cell(row = br+1 ,column = i).value = v
                              io.cell(row = br+1 ,column = i).font = copy(io.cell(row = br,column = i).font)
                              io.cell(row = br+1 ,column = i).border = copy(io.cell(row = br,column = i).border)
                              io.cell(row = br+1 ,column = i).fill = copy(io.cell(row = br,column = i).fill)
                            br = br + 1;
                              #print(io.cell(row = 6,column = 4).value);
                            return br
                        
                        row = blank_row-1;
                        
                         
                        
                                               
                        
                     
                         
                        for index in range(0,len(consumption_dict[Reg]['exchanges'])):
                            
                            
                            blank_row = createblnkrow(blank_row)
                            io[row][0].value = index+1;
                            
                            
                            
                            if index == 0:
                               io[row][2].value = 0;
                               io[row][3].value = consumption_dict[Reg]['name']
                            else:
                                io[row][1].value = 5;
                                io[row][3].value = consumption_dict[Reg]['exchanges'][index]['flow']['name']
                            
                            io[row][4].value = consumption_dict[Reg]['category']
                            if len(consumption_dict[Reg]['exchanges']) == 2:
                               io[row][5].value = str(Reg)
                            else:
                               if index == 1: 
                                  io[row][5].value = consumption_dict[Reg]['exchanges'][index]['flow']['name'][23:27]
                               else:
                                  io[row][5].value = str(Reg)
                            io[row][6].value = consumption_dict[Reg]['exchanges'][index]['amount']
                            io[row][7].value = 'MWh';
                            io[row][21].value = consumption_dict[Reg]['description']                         
                            row = row+1;
                            index=index+1
                            
                        
                                                     

                        
                                        
                             
                        filename = Reg+"_Consumption.xlsx"
                        data_dir = os.path.dirname(os.path.realpath(__file__))+"/output/"
                        wb.save(data_dir+filename)
                        print(filename+' File written Successfully')




def surplus_pool_mix_template_generator(surplus_pool_dict,nerc_region):

        data_dir = os.path.dirname(os.path.realpath(__file__))   
        os.chdir(data_dir+'/data/')            
         
        nerc_region = [list(i)[0].value for i in nerc_region]
    
        for Reg in nerc_region:
                        
                            
                        wb = openpyxl.load_workbook('template.xlsx')
                        
                        gi = wb['General information']
                        io = wb['InputsOutputs']
                        
                        
                        global blank_row;
                                   
                        blank_row = 6;
                        Column_number = 43 #THis is based on the template number of columns in the template. 
                           
                        
                        index = 1; #Starting the index at a specific point
                        #REading the list of fuel for the entered region
                        
                        
                                       
                        
                        #This block is used for filling up the General Infomration Sheet
                        
                        
                                
                        
                        
                        
                        gi['D4'].value = 'Electricity '
                        
                        gi['D5'].value = 'for surplus_pool';
                        
                        gi['D6'].value = 'at region '+ Reg;
                        
                        gi['D7'].value = 'Surplus Pool';
                        
                        gi['D10'].value = 'surplus_pool mix electricity in the '+Reg+' region'
                        
                        
                        gi['D11'].value = surplus_pool_dict[Reg]['category']
                        
                        gi['D12'].value = 'FALSE'
                        
                        gi['D14'].value = 'Electricity; voltage?'
                        
                        gi['D16'].value = surplus_pool_dict[Reg]['processDocumentation']['validFrom']
                        
                        gi['D17'].value = surplus_pool_dict[Reg]['processDocumentation']['validUntil']
                        
                        gi['D18'].value = surplus_pool_dict[Reg]['processDocumentation']['validFrom']
                        
                        gi['D20'].value = 'US-eGRID-'+Reg
                        
                        v= '' ;
                        #THis function is used for finding the NERC regions. THese input files are built manually. 
            
                        
                        gi['D24'].value = 'EGRID subregion definitions:<https://www.epa.gov/energy/emissions-generation-resource-integrated-database-egrid>\nNERC region: \nStates totally or partially included: <autofill>'
                        
                        #This is for printing the Prime Movers. Written to a output file and reread back. 
                        #def pm(fn,rg):
                        #    f = open('out.txt', 'w')
                        #    for row in primemover.itertuples():
                        ##      if row[2] == rg:
                         #       if row[3]  == fn or row[4] == fn or row[5] == fn or row[6] == fn:
                        #          if row[7] != None:                   
                        #             f.write(row[7])
                        #             f.write(',')
                        #    f.close()        
                        
                         
                        
                        
                        #linestring = open('out.txt', 'r').read()
                        
                        gi['D26'].value = 'This is the surplus_pool Mix.'
                        
                
                        
                        
                        #This block is used for filling up the InputOutput Sheet ONLY EMISSIONS
                        
                                               
                        
                        
                        #This function is necessary to creates new line in the template for writing input and output flows. 
                        def createblnkrow(br):
                            for i in range(1,Column_number):
                              
                              v = io.cell(row = br,column = i).value
                              io.cell(row = br+1 ,column = i).value = v
                              io.cell(row = br+1 ,column = i).font = copy(io.cell(row = br,column = i).font)
                              io.cell(row = br+1 ,column = i).border = copy(io.cell(row = br,column = i).border)
                              io.cell(row = br+1 ,column = i).fill = copy(io.cell(row = br,column = i).fill)
                            br = br + 1;
                              #print(io.cell(row = 6,column = 4).value);
                            return br
                        
                        row = blank_row-1;
                        
                         
                        
                                               
                        
                     
                         
                        for index in range(0,len(surplus_pool_dict[Reg]['exchanges'])):
                            
                            
                            blank_row = createblnkrow(blank_row)
                            io[row][0].value = index+1;
                            
                            if index == 0:
                                io[row][2].value = 0;
                                io[row][3].value = surplus_pool_dict[Reg]['name']
                            else:
                                io[row][1].value = 5;
                                io[row][3].value = surplus_pool_dict[Reg]['exchanges'][index]['flow']['name']
                            
                            io[row][4].value = surplus_pool_dict[Reg]['category']
                            if index == 0:
                                io[row][5].value = str(Reg)
                            else:
                                io[row][5].value = surplus_pool_dict[Reg]['exchanges'][index]['location']
                            io[row][6].value = surplus_pool_dict[Reg]['exchanges'][index]['amount']
                            io[row][7].value = 'MWh';
                            io[row][21].value = surplus_pool_dict[Reg]['description']                         
                            row = row+1;
                            index=index+1
                            
                        
                                                     

                        
                                        
                             
                        filename = Reg+"_surplus_pool.xlsx"
                        data_dir = os.path.dirname(os.path.realpath(__file__))+"/output/"
                        wb.save(data_dir+filename)
                        print(filename+' File written Successfully')
                        
                        
                        
def trade_mix_template_generator(trade_dict):
    
        data_dir = os.path.dirname(os.path.realpath(__file__))   
        os.chdir(data_dir+'/data/')            
         
        
    
        for Reg in trade_dict.keys():
                        
                            
                        wb = openpyxl.load_workbook('template.xlsx')
                        
                        gi = wb['General information']
                        io = wb['InputsOutputs']
                        
                        
                        global blank_row;
                                   
                        blank_row = 6;
                        Column_number = 43 #THis is based on the template number of columns in the template. 
                           
                        
                        index = 1; #Starting the index at a specific point
                        #REading the list of fuel for the entered region
                        
                        
                                       
                        
                        #This block is used for filling up the General Infomration Sheet
                        
                        
                                
                        
                        
                        
                        gi['D4'].value = 'Electricity '
                        
                        gi['D5'].value = 'for trade';
                        
                        gi['D6'].value = 'at region '+ Reg;
                        
                        gi['D7'].value = 'trade Mix';
                        
                        gi['D10'].value = 'Trade mix electricity in the '+Reg+' region'
                        
                        
                        gi['D11'].value = trade_dict[Reg]['category']
                        
                        gi['D12'].value = 'FALSE'
                        
                        gi['D14'].value = 'Electricity; voltage?'
                        
                        gi['D16'].value = trade_dict[Reg]['processDocumentation']['validFrom']
                        
                        gi['D17'].value = trade_dict[Reg]['processDocumentation']['validUntil']
                        
                        gi['D18'].value = trade_dict[Reg]['processDocumentation']['validFrom']
                        
                        gi['D20'].value = Reg
                        
                        v= '' ;
                        #THis function is used for finding the NERC regions. THese input files are built manually. 
            
                        
                        gi['D24'].value = 'Trade from Outside US'
                        #This is for printing the Prime Movers. Written to a output file and reread back. 
                        #def pm(fn,rg):
                        #    f = open('out.txt', 'w')
                        #    for row in primemover.itertuples():
                        ##      if row[2] == rg:
                         #       if row[3]  == fn or row[4] == fn or row[5] == fn or row[6] == fn:
                        #          if row[7] != None:                   
                        #             f.write(row[7])
                        #             f.write(',')
                        #    f.close()        
                        
                         
                        
                        
                        #linestring = open('out.txt', 'r').read()
                        
                        gi['D26'].value = 'This is the trade Mix.'
                        
                        filename = 'trade Mix '+ Reg + '.xlsx'
                        
                        
                        
                        #This block is used for filling up the InputOutput Sheet ONLY EMISSIONS
                        
                                               
                        
                        
                        #Reading All the files
                        #io['D5'].value =  Reg+' Distribution Mix electricity';   
                        
                        #This function is necessary to creates new line in the template for writing input and output flows. 
                        def createblnkrow(br):
                            for i in range(1,Column_number):
                              
                              v = io.cell(row = br,column = i).value
                              io.cell(row = br+1 ,column = i).value = v
                              io.cell(row = br+1 ,column = i).font = copy(io.cell(row = br,column = i).font)
                              io.cell(row = br+1 ,column = i).border = copy(io.cell(row = br,column = i).border)
                              io.cell(row = br+1 ,column = i).fill = copy(io.cell(row = br,column = i).fill)
                            br = br + 1;
                              #print(io.cell(row = 6,column = 4).value);
                            return br
                        
                        row = blank_row-1;
                        
                         
                        
                                               
                        
                     
                         
                        for index in range(0,len(trade_dict[Reg]['exchanges'])):
                            
                            
                            blank_row = createblnkrow(blank_row)
                            io[row][0].value = index+1;
                            
                            
                            
                            if index == 0:
                               io[row][2].value = 0;
                               io[row][3].value = trade_dict[Reg]['name']
                            else:
                                io[row][1].value = 5;
                                io[row][3].value = trade_dict[Reg]['exchanges'][index]['flow']['name']
                            
                            io[row][4].value = trade_dict[Reg]['category']
                            io[row][5].value = trade_dict[Reg]['location']
                            
                            io[row][6].value = trade_dict[Reg]['exchanges'][index]['amount']
                            io[row][7].value = 'MWh';
                            io[row][21].value = trade_dict[Reg]['description']                         
                            row = row+1;
                            index=index+1
                            
                        
                                                     
                         
                                        
                             
                        filename = Reg+"_trade.xlsx"
                        data_dir = os.path.dirname(os.path.realpath(__file__))+"/output/"
                        wb.save(data_dir+filename)
                        print(filename+' File written Successfully')
