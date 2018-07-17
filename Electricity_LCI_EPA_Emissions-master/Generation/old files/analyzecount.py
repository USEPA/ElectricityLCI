# -*- coding: utf-8 -*-
"""
Created on Thu May 24 19:10:01 2018

@author: TGhosh
"""

import pandas as pd
from copy import copy, deepcopy
import openpyxl
import numpy as np
import os
import sys

def egrid_func(a,b):

        #Reading the facility file
        egrid1 = pd.read_csv("egrid_2016_1.csv", header=0, error_bad_lines=False)
        
        
        #Reading the flow by facility file
        egrid2 = pd.read_csv("egrid_2016.csv", header=0, error_bad_lines=False)
        
        
        egrid3 = egrid2.pivot(index = 'FacilityID',columns = 'FlowName', values = 'FlowAmount').reset_index()
        
        
        #Merging dataframes together
        egrid = egrid1.merge(egrid3, left_on = 'FacilityID', right_on = 'FacilityID')
        egrid[['Heat input','Net generation']] = egrid[['Heat input','Net generation']].apply(pd.to_numeric,errors = 'coerce')
        
        
        
        #Calculating efficiency
        egrid.loc[:,"Efficiency"] = egrid['Net generation'].values*100/egrid['Heat input'].values
        egrid = egrid.sort_values(by = ['Efficiency'], axis=0, ascending=True, inplace=False, kind='quicksort', na_position='last')
       
        #Replace inf
        egrid = egrid.replace([np.inf, -np.inf], np.nan)
        #Replacing all the facilities with no net generation or Heat input reported
        egrid = egrid.dropna(subset = ['Efficiency'])
       
        #Dropping less than 10%
        egrid_new1 = egrid[(egrid['Efficiency'] >= a) & (egrid['Efficiency'] <= b)]
        
        
        
       
        
        col = list(egrid_new1['Heat input'])        
        egrid_new1['HeatInput'] = col;
        cols = egrid_new1.columns.tolist()
        egrid_new1 = egrid_new1[[cols[-1]] + cols[:-1]] 
        
        
        col = list(egrid_new1['Net generation'])         
        egrid_new1['NetGen'] = col;
        cols = egrid_new1.columns.tolist()
        egrid_new1 = egrid_new1[[cols[-1]] + cols[:-1]] 
        egrid_new1 = egrid_new1.drop(columns = ['Net generation'])       
        egrid_new1 = egrid_new1.drop(['Heat input','Efficiency'],axis = 1)
        
        
        return egrid_new1
    
    
    
    
    
def tri_func(egrid):
        
        #READING tri database
        tri = pd.read_csv("TRI_2016.csv", header=0, error_bad_lines=False)  
        
        #Droppping duplicates in STEWI
        tri = tri.drop_duplicates(keep = 'first')
        tri = tri[['FacilityID','FlowName','Compartment','FlowAmount']]
        tri2 = tri.groupby(['FacilityID','FlowName','Compartment'])['FlowAmount'].sum()
        tri2 = tri2.reset_index()
        
        #tri2 = tri2.dropna(axis = 0, how = 'any')
        #TRI reshaping pivot not working so using pivot table
        tri3 = pd.pivot_table(tri2,index = ['FacilityID','Compartment'], columns = 'FlowName', values = 'FlowAmount')
        
        tri3 = tri3.reset_index()
        #print(tri2[['ReleaseType']])
        
        tri2egrid = pd.read_csv("tri_egrid.csv", usecols = ['EGRID_ID','TRI_ID1'],header=0, error_bad_lines=False)
        
        #Merging with the bridge file with EGRID
        tri4 = tri2egrid.merge(tri3, left_on = 'TRI_ID1', right_on = 'FacilityID')
        tri4 = tri4.drop(columns = ['FacilityID','TRI_ID1'])
        
        
        egrid = egrid[['NetGen','FacilityID']]
                
        #Merging egrid and TRI
        database = egrid.merge(tri4,left_on ='FacilityID', right_on = 'EGRID_ID')
        
        database = database.drop(columns = ['EGRID_ID','NetGen','Compartment'])
        return database   



def main(a,b):
        wbook = openpyxl.load_workbook('Region.xlsx')
        wb = wbook['Sheet2']
        fuel_name = pd.read_csv("fuelname.csv", header=0, error_bad_lines=False)
        reg_name = pd.read_excel("Region.xlsx", header=None, error_bad_lines=False)
        i = 0;
        for row in reg_name.itertuples():
            i = i+1;
            database = egrid_func(a,b)
            
            database = database[database['eGRID subregion acronym'] == str(row[1])]
            #print(database['eGRID subregion acronym'])
            
            
            for roww in fuel_name.itertuples():
                            #croppping the database according to the current fuel being considered
                    database_f1 = database[database['Plant primary coal/oil/gas/ other fossil fuel category'] == roww[1]]
                    
                    if database_f1.empty == True:
                          database_f1 = database[database['Plant primary fuel'] == roww[1]]              
                    tri_db = tri_func(database_f1)
                    database_f3 = database_f1[['Carbon dioxide', 'Nitrous oxide', 'Nitrogen oxides', 'Sulfur dioxide', 'Methane']]
                    
                    k=0;
                    
                    wb[i+1][0].value = str(row[1])+' '+str(roww[1]) ;                 
                    #for j in database_f3.iteritems():
                    #   k = k+1;
                    #   wb.cell(row = 1, column = k+1).value = j[0];
                    #   wb.cell(row = i+1, column = k+1).value = database_f3[j[0]].count();
                    i = i+1;
                    k = k+1;
                    wb.cell(row = 1, column = k+1).value = 'eGRID'
                    wb.cell(row = 1, column = k+2).value = 'TRI'
                    wb.cell(row = i+1, column = k+1).value = database_f1['FacilityID'].count();
                    wb.cell(row = i+1, column = k+2).value = tri_db['FacilityID'].count();
                    #for u in tri_db.iteritems():
                    #   k = k+1;
                    #   wb.cell(row = 1, column = k+1).value = u[0];
                    #   wb.cell(row = i, column = k+1).value = tri_db[u[0]].count();                  
        
        
        
        wbook.save('count.xlsx')
    
z1 = main(10,100)           