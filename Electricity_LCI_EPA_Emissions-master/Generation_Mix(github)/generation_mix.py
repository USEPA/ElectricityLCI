#Writes out generation mixes. Not currently working because relies on inventory from old sources

import pandas as pd
from copy import copy, deepcopy
import openpyxl
import numpy as np
import os
import sys

def egrid_func(a,b):

        #Reading the facility file
        egrid1 = pd.read_csv("egrid_2014_1.csv", header=0, error_bad_lines=False)
        
        
        #Reading the flow by facility file
        egrid2 = pd.read_csv("egrid2014_trircrainfoneiegrid.csv", header=0, error_bad_lines=False)
        
        egrid2 = egrid2[egrid2['Source'] == 'eGRID']
        
        egrid2 = egrid2.drop_duplicates(keep = 'first')
        
        
        egrid3 = pd.pivot_table(egrid2,index = 'FacilityID' ,columns = 'FlowName', values = 'FlowAmount').reset_index()
        
        
        #Merging dataframes together
        egrid = egrid1.merge(egrid3, left_on = 'FacilityID', right_on = 'FacilityID')
        egrid[['Heat input','Net generation']] = egrid[['Heat input','Net generation']].apply(pd.to_numeric,errors = 'coerce')
        
        
        
        #Calculating efficiency
        egrid.loc[:,"Efficiency"] = egrid['Net generation'].values*100/egrid['Heat input'].values
        egrid = egrid.sort_values(by = ['Efficiency'], axis=0, ascending=True, inplace=False, kind='quicksort', na_position='last')
       
        #Replace inf
        egrid = egrid.replace([np.inf, -np.inf], np.nan)
        #Replacing all the facilities with no net generation or Heat input reported
        egrid = egrid.dropna(subset = ['Efficiency'])
       
        #Dropping less than 10%
        egrid_new1 = egrid[(egrid['Efficiency'] >= a) & (egrid['Efficiency'] <= b)]
        
        
        
       
        
        col = list(egrid_new1['Heat input'])        
        egrid_new1['HeatInput'] = col;
        cols = egrid_new1.columns.tolist()
        egrid_new1 = egrid_new1[[cols[-1]] + cols[:-1]] 
        
        
        col = list(egrid_new1['Net generation'])         
        egrid_new1['NetGen'] = col;
        cols = egrid_new1.columns.tolist()
        egrid_new1 = egrid_new1[[cols[-1]] + cols[:-1]] 
        egrid_new1 = egrid_new1.drop(columns = ['Net generation'])       
        egrid_new1 = egrid_new1.drop(['Heat input','Efficiency'],axis = 1)
        egrid_new1['NetGen'] = egrid_new1['NetGen']*0.277778/1000
        #egrid_new1.to_csv('chk.csv')
        return egrid_new1
 
    
    
    
#This function is necessary to creates new line in the template for writing input and output flows. 
def createblnkrow(br,io):
    
    Column_number = 43 #THis is based on the template number of columns in the template. 
    for i in range(1,Column_number):
      
      v = io.cell(row = br,column = i).value
      io.cell(row = br+1 ,column = i).value = v
      io.cell(row = br+1 ,column = i).font = copy(io.cell(row = br,column = i).font)
      io.cell(row = br+1 ,column = i).border = copy(io.cell(row = br,column = i).border)
      io.cell(row = br+1 ,column = i).fill = copy(io.cell(row = br,column = i).fill)
    
    br = br + 1;
      #print(io.cell(row = 6,column = 4).value);
    return br


    
def generator(a,b,Reg):
            
          global year; 
          year = '2014'
          database = egrid_func(a,b)
          database = database[database['eGRID subregion acronym'] == Reg]
        
          wbook = openpyxl.load_workbook('template.xlsx')                
        
          fuel_name = pd.read_csv("fuelname.csv", header=0, error_bad_lines=False)

              #assuming the starting row to fill in the emissions is the 5th row Blank
              
          global blank_row;    
          blank_row = 6;
          index = 2;
          
                                  
          names = pd.read_excel('Name of flows.xlsx')
        #Read the template files.                         
          
          gi = wbook['General information']
        
        #This block is used for filling up the General Infomration Sheet      
          gi['D4'].value = 'Electricity'
        
          gi['D5'].value = 'from Generation';
        
          gi['D6'].value = 'at region '+Reg;
        
        #Mix type
          gi['D7'].value = 'Production Mix'; 
        
          gi['D9'].value = 'Electricity; from Generation mix; at region '+Reg; 
        
          gi['D10'].value = 'Electricity from generation mix using power plants in the '+Reg+' region'
        
        #This function is used for the name of the fuels.
        #def pm1(fn):
        #    for row in primemover.itertuples():
        #       if row[3]  == fn or row[4] == fn or row[5] == fn or row[6] == fn:
        #          if row[4] != None:                   
        #             return row[4].capitalize()
        
          gi['D11'].value = '22: Utilities/2211: Electric Power Generation, Transmission and Distribution/'
        
          gi['D12'].value = 'FALSE'
        
          gi['D14'].value = 'Electricity; voltage?'
        
          gi['D16'].value = '01/01/2017'
        
          gi['D17'].value = '12/31/2017'
        
          gi['D18'].value = year
        
          gi['D20'].value = 'US-eGRID-'+Reg
    
          gi['D21'].value = 'F'
          v= [];
          v1= [];
        
          for roww in database.itertuples():
               
              if Reg == roww[5]:
                  v1.append(str(roww[4]))
                  v.append(str(roww[11]))
                
          v2 = list(set(v1)) 
        
          str1 = ','.join(v2)
        
          v2 = list(set(v))
          
          str2 = ','.join(v2)
        
          gi['D24'].value = 'database subregion definitions:<https://www.epa.gov/energy/emissions-generation-resource-integrated-database-database>\nNERC region: '+str2+'\nStates totally or partially included: '+str1
       
          gi['D26'].value = 'This is an aggregation of technology types within this EGRID subregion'
          
                    
          for row in fuel_name.itertuples():
                
                fuelname = row[2]
                print(fuelname)
                #The fuels and their types are in two different columns
    
                    
                #croppping the database according to the current fuel being considered
                database_f1 = database[database['Plant primary coal/oil/gas/ other fossil fuel category'] == row[1]]
                if database_f1.empty == True:
                      database_f1 = database[database['Plant primary fuel'] == row[1]]              
                
                if database_f1.empty == False:
              
                    #This block is used for filling up the InputOutput Sheet ONLY EMISSIONS
                    io = wbook['InputsOutputs']
                    
                    
        
                    #Fillin up with reference only only
                    io['A5'].value =  1
                    io['C5'].value =  0;
                    io['D5'].value =  str(names.iloc[0,0])
                    io['E5'].value = '22: Utilities/2211: Electric Power Generation, Transmission and Distribution/'
                    io['F5'].value = str(Reg)
                    io['H5'].value =  'MWh'
                    row1 = blank_row;
                    
                               
                    
        
                    #Fillin up the fuel flow
                    blank_row = createblnkrow(blank_row,io)
                    io[row1][0].value = index;
                    io[row1][1].value = 5;
                    #io[row1][3].value = 'Electricity from '+fuelname+' at '+Reg;
                    io[row1][3].value = str(names.iloc[0,0])+' '+fuelname
                    #io[row1][4].value = '22: Utilities/2211: Electric Power Generation, Transmission and Distribution/'+fuelname;
                    io[row1][5].value = Reg
                    io[row1][6].value = np.sum(database_f1['NetGen'])/np.sum((database['NetGen']))
                    io[row1][7].value = 'MWh';
                    io[row1][9].value = 'Electricity; from '+ fuelname+'; at generating facility';
                    io[row1][21].value = 'database data with plants over 10% efficiency';
                    #io[row1][26].value = 'Normal';
                    io[row1][27].value = np.sum(database_f1['NetGen'])/np.sum((database['NetGen']))
                    #io[row1][28].value = database_f1[i[0]].std();
                    #io[row1][29].value = database_f1[i[0]].min();
                    #io[row1][30].value = database_f1[i[0]].max();
                    io[row1][33].value = 'eGRID '+year;
                    
                    row1 = row1+1;
                    index = index+1;
                    
                  
                

          filename = Reg+"_Mix.xlsx"
          data_dir = os.path.dirname(os.path.realpath(__file__))+"\\results\\"
          wbook.save(data_dir+filename)
                




p = egrid_func(10,100)

region = pd.read_excel('Region.xlsx')
for ind in region.itertuples():
    print(ind[1])
    
    generator(10,100,str(ind[1]))