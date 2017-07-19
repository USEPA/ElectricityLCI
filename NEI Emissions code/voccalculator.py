# -*- coding: utf-8 -*-
"""
Created on Tue Jul 18 12:07:17 2017

@author: tghosh
"""

import numpy as np
#import plotly.plotly as py
import openpyxl
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
import os
import statistics

print('Opening workbook...')


#NEI name
wb = openpyxl.load_workbook('VOCnamematched.xlsx')
chk = wb.get_sheet_by_name('Sheet1')

vocname = chk['F2':'F172']


wb1 = openpyxl.load_workbook('FINALNEIEMISSIONSnamechngd.xlsx')
chkk = wb1.get_sheet_by_name('finalNEIemissions')

neiname = chkk['G1':'HW1']
voc = chkk['HL2':'HL2219']

for i in range(0,2218):
    sum = 0;
    if voc[i][0].value != None:
       for j in range(0,225):
           for k in range(0,171):
               if neiname[0][j].value == vocname[k][0].value:
                   if(chkk.cell(row = i+2,column = j+7).value != None):
                       sum = sum + chkk.cell(row = i+2,column = j+7).value;
       chkk.cell(row = i+2, column = 233).value = sum;
       
       
       
wb1.save('finalemissionsvoc_name_1.xlsx')       