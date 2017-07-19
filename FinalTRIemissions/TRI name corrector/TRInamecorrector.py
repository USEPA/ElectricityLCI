# -*- coding: utf-8 -*-
"""
Created on Wed Jul 19 11:25:09 2017

@author: tghosh
"""


import numpy as np
#import plotly.plotly as py
import openpyxl
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
import os
import statistics

print('Opening workbook...')

wb = openpyxl.load_workbook('Namebridgesoil.xlsx')
chk = wb.get_sheet_by_name('Sheet1')

oldname = chk['C2':'C471']
newname = chk['H2':'H471']

del chk


wb1 = openpyxl.load_workbook('chk.xlsx')
chk = wb1.get_sheet_by_name('Sheet1')



wb2 = openpyxl.load_workbook('TRI.xlsx')
sheet = wb2.get_sheet_by_name('TotalLandTreat')

oldname1 = sheet['G1':'IL1']

del sheet

redFill = PatternFill(start_color='FFF7A5E4',
                   end_color='FFF7A5E4',
                   fill_type='solid')

blueFill = PatternFill(start_color='FFB8A1F8',
                   end_color='FFB8A1F8',
                   fill_type='solid')

for i in range(0,240):
    s = 0;
    for j in range(0,470):        
        chk.cell(row = 1,column = i+1).value = oldname1[0][i].value;
        if oldname[j][0].value == oldname1[0][i].value:
            s = 1;
            v = newname[j][0].value
            if v == None:
               chk.cell(row = 2,column = i+1).fill = redFill
               v2 = oldname[j][0].value;
               chk.cell(row = 2,column = i+1).value = v2;
               
            else:   
               chk.cell(row = 2,column = i+1).value = v;
            
            
    if s==0:
           chk.cell(row = 2,column = i+1).value = oldname1[0][i].value;
           chk.cell(row = 2,column = i+1).fill = blueFill
           s=0;

wb1.save('Temp.xlsx')
             
               
'''            





wb2 = openpyxl.load_workbook('TRI.xlsx')
sheet = wb2.get_sheet_by_name('StreamDischgE')

oldname1 = sheet['G1':'IL1']

del sheet

redFill = PatternFill(start_color='FFF7A5E4',
                   end_color='FFF7A5E4',
                   fill_type='solid')

for i in range(0,240):
    for j in range(0,470):
        chk.cell(row = 1,column = i+1).value = oldname1[0][i].value;
        if oldname[j][0].value == oldname1[0][i].value:
            v = newname[j][0].value
            if v == None:
               chk.cell(row = 2,column = i+1).fill = redFill
               v2 = oldname[j][0].value;
               chk.cell(row = 2,column = i+1).value = v2;
               
            else:   
               chk.cell(row = 2,column = i+1).value = v;
            
            
        #else:
         #   if oldname1[0][i].value != None:
          #     chk.cell(row = 2,column = i+1).value = oldname1[0][i].value;
               
               
      

wb1.save('Temp.xlsx')

'''